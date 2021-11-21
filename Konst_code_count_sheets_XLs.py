import sys
import os
from time import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
    
def set_files():
    global write_path, read_file, wb, sheets, sheet, xl_files, rowcount, ac_row
    if len(sys.argv) > 1:
        read_file = sys.argv[1]
        if read_file[-4:].find(".xlsx") != 0:
            raise Exception("Δεν είναι αρχείο xlsx αυτό...")
    else:
        write_path = ("Z:\\OneDrive\\eShop Stuff\\Synced\\Κουτσοδούλια\\KonstaStock\\11-2021")
        os.chdir(write_path)
        read_file = "2016_code_analysis.xlsx"
    
    wb = load_workbook(read_file, data_only = True)
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    
    rowcount = sheet.max_row
    ac_row = 0
    for i in range(1, rowcount):
        row_value = str(sheet.cell(row=i, column=1).value)
        if row_value == "" or row_value is None or row_value == "None":
            break
        else:
            ac_row += 1


try:
    set_files()
    code_list = []
    code_quant = []
    code_final_price = []
    quantity = 0
    final_price = 0
    for i in range(2, ac_row + 1):
        cur_code = str(sheet.cell(row=i, column=2).value)
        code_list.append(cur_code)
    code_set = set(code_list)
    code_list = list(code_set)
    for index, code in enumerate(code_list):
        for c in range(1, ac_row):
            cur_code = str(sheet.cell(row=c, column=2).value)
            if cur_code == code:
                cur_quant = int(sheet.cell(row=c, column=4).value)
                quantity += cur_quant
                if len(code_quant) == 0:
                    code_quant.append(quantity)
                else:
                    code_quant[index] += quantity
                cur_price = round(sheet.cell(row=c, column=6).value, 2)
                final_price = round(cur_quant * cur_price, 2)
                if len(code_final_price) == 0:
                    code_final_price.append(final_price)
                else:
                    code_final_price[index] += final_price
            print("index:", index)
            print("code:", code)
            print("code_quant:", code_quant[index])
            print("code_final_price:", code_final_price[index])
except KeyboardInterrupt:
    print("bye")
except Exception as exc:
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print("Exception: " + str(exc))
    print("Exception type: ", exception_type)
    print("File name: ", filename)
    print("Line number: ", line_number)
