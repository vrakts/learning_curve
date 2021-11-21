# code_seperation.py

import sys
import os
from time import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
    

def get_started():
    global start_time
    start_time = time()


def get_finished():
    global finish_time, formatted_time
    finish_time = time()
    diff = finish_time - start_time
    mins = int(diff / 60)
    hours = int(mins / 60)
    seconds = int(diff - (mins * 60))
    formatted_time = str(hours).zfill(2) + ":" + str(mins).zfill(2) + ":" + str(seconds).zfill(2)


def set_files():
    global write_path, read_file, wb, sheets, xl_files
    if len(sys.argv) > 1:
        read_file = sys.argv[1]
        if read_file[-4:].find(".xlsx") != 0:
            raise Exception("Δεν είναι αρχείο xlsx αυτό...")
    else:
        write_path = ("Z:\\OneDrive\\eShop Stuff\\Synced\\Κουτσοδούλια\\KonstaStock\\11-2021")
        os.chdir(write_path)
        read_file = "2016-2017_web.xlsx"
    
    wb = load_workbook(read_file)
    sheets = wb.sheetnames
    xl_files  = []
    for name in sheets:
        xl_files.append("code_analysis_" + name + ".xlsx")
    

def count_sheet(sheet):
    global colcount, ac_col, col_index, rowcount, ac_row, row_index, wb_write, ws_write
    wb_write = Workbook()
    ws_write = wb_write.active
    ws_write.title = "codes"
    ws_write.cell(row=1, column=1, value="ΠΑΡΑΓΓΕΛΙΑ")
    ws_write.cell(row=1, column=2, value="ΚΩΔΙΚΟΣ")
    ws_write.cell(row=1, column=3, value="ΤΙΤΛΟΣ")
    ws_write.cell(row=1, column=4, value="ΤΕΜΑΧΙΑ")
    ws_write.cell(row=1, column=5, value="ΤΙΜΗ")
    ws_write.column_dimensions['A'].width = 24
    ws_write.column_dimensions['B'].width = 12
    ws_write.column_dimensions['C'].width = 24
    ws_write.column_dimensions['D'].width = 10
    ws_write.column_dimensions['E'].width = 10

    colcount = sheet.max_column
    ac_col = 0
    for i in range(1, colcount):
        col_value = str(sheet.cell(row=1, column=i).value)
        if col_value == "" or col_value is None or col_value == "None":
            break
        else:
            ac_col += 1

    col_index = 2
    # Μέτρημα και επιλογή γραμμών
    rowcount = sheet.max_row
    ac_row = 0
    for i in range(1, rowcount):
        row_value = str(sheet.cell(row=i, column=1).value)
        if row_value == "" or row_value is None or row_value == "None":
            break
        else:
            ac_row += 1
    row_index = 1


def fill_list(temp_text):
    global code_list
    code_list = []
    for t in temp_text.split("|"):
        temp_item = t.replace("|", "").strip()
        # print(temp_item)
        code_list.append(temp_item)


try:
    set_files()
    cur_row = 2
    cur_col = 2
    index = 0
    get_started()
    for s, sheet in enumerate(sheets):
        sheet = wb[sheet]
        xl_file = xl_files[s]
        count_sheet(sheet)
        for r in range(row_index, ac_row + 1):
            cell_value = str(sheet.cell(row=r, column=col_index).value)
            temp_text = cell_value.replace("][", "|").replace("[", "").replace("]", "")
            fill_list(temp_text)        
            
            print("")
            
            for i in range(0, len(code_list)):
                item = code_list[i]
                order = str(sheet.cell(row=r, column=1).value)
                # print("r:", r)
                print("order:", order + ", item:", item)
                """ print("i:", i)
                print("item:", item)
                print("index:", index)
                print("index % 3:", index % 3)
                print("index % 5:", index % 5) """

                if index > 1:
                    if index % 3 == 0:
                        #input()
                        index += 1
                        continue
                    elif index % 4 == 0:
                        print("")
                    elif index % 5 == 0:
                        index = 0
                        cur_row += 1
                        cur_col = 2
                    
                ws_write.cell(row=cur_row, column=1, value=str(sheet.cell(row=r, column=1).value))
                ws_write.cell(row=cur_row, column=cur_col, value=item)
                index += 1
                cur_col += 1
        wb_write.save(xl_file)
except KeyboardInterrupt:
    print("φρένο")
    # for item in code_list:
    #     print(item)
except Exception as exc:
    print("exception:", str(exc))

# wb_write.save(xl_file)
get_finished()
print(formatted_time)