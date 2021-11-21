# links_test.py
import openpyxl
print("Loading excel file...")
# wb = openpyxl.load_workbook(r'Z:\times.xlsx')
wb = openpyxl.load_workbook(r'Z:\OneDrive\HTML Parser\Python\Αλλαγή τιμών.xlsx')
print("Done.")
print("")
ws = wb['SSD']
print("Selected SSD")
print("")
# print(ws.cell(row=4, column=9).value)
# print(ws.cell(row=4, column=9).hyperlink.target)
row_count = ws.max_row
column_count = ws.max_column

for i in range(1, row_count) :
 if(ws.cell(row=i, column=1).value) == "" :
  print("Row: " + str(i) + ", Column: 1 = Empty")
  break
 else :
  print("Row: " + str(i))
  for j in range(1, 15) :
   print("Row: " + str(i) + ", Column: " + str(j))
   try :
    print(ws.cell(row=i, column=j).value)
    print(ws.cell(row=i, column=j).hyperlink.target)
   except :
    print("No link found")
  print("")
  input()
