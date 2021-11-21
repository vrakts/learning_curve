import sys  # for exit purposes in case of error


def ti_paizei():
    version = "1.6 Beta"
    # Current version 1.6 beta
    # Ελέγχει αν υπάρχει μετάφραση στα αγγλικά στα προϊόντα συγκεκριμένης κατηγορίας.
    # Διαβάζει από ένα αρχείο τα links ή τα ζητάει από τον χρήστη αν δεν βρεθεί.
    # Αν δεν δωθεί κάποιο switch στην εκκίνηση αναζητάει μόνο στο αγγλικό site αν υπάρχει μετάφραση.
    # Με το switch -el κοιτάει και το ελληνικό site αν δεν υπάρχει ελληνική περιγραφή.
    # Με το switch -war ελέγχει το κείμενο της εγγύησης και το κατγράφει και αυτό.
    # Προσοχή: Με το switch war καταγράφονται όλα τα αποτελέσματα ακόμα και αν δεν λείπει περιγραφή / μετάφραση.
    #####################
    # Changelog v1.6
    # - Καλύτερα definitions ιδίως στο load_soup.
    # - Αν υπάρχει το αρχείο διαβάζει από εκεί αλλιώς ζητάει URL
    # - Σημειώνει αν έχει ελληνική ή αγλλική μετάφραση.
    # - Γράφει και το κείμενο της εγγύησης στα αποτελέσματα.
    # - Multi process για το φόρτωμα σούπας ελληνικά / αγγλικά.
    # - Μικροδιορθώσεις στο παραπάνω για το λάθος στο export => .exe (freeze_support())
    # - Δοκιμή για MultiThreading αντί MultiProcessing
    # - Νέος κώδικας για μέτρηση χρόνου
    # - Πληροφορίες στον τίτλο
    # - Προσπάθεια για
    #      Αναγνώριση των Crazy προϊόντων από την φωτογραφία που θα υπάρχει μέσα στη σελίδα. - Seems OK
    #####################
    # Changelog v1.5.1
    # - Γράφει και τον τίτλο στο excel.
    #####################
    # Changelog v1.4
    # - Υπολογίζει και τα Crazy προϊόντα και γράφει την αντίστοιχη σήμανση στο excel.
    # - Διορθώσεις στο for loop για καλύτερες επιδόσεις.
    #####################
    # Changelog v1.3
    # - Μαζεύει μπόλικα link και τα τρέχει ένα ένα.
    #####################
    # Changelog v1.2
    # - Τώρα το πρόγραμμα μιλάει στα Ελληνικά.
    # - Теперь программа говорит по-русски, вероятно, неправильно, хотя.
    #####################
    # Changelog v1.1
    # - Added a test write path for the Manager's PC
    # - Enclosed all processes in functions for easy calling.
    # - Try and Except for various errors
    # - Cosmetic changes.
    #####################
    # Changelog v1
    # - From a category URL check all product descriptions for EN translation.
    # - Writes non translated to a file.
    #####################
    # To Do
    # - Προσθήκη κάτι σαν το create URL να μαζεύει όλα τα found
    #   σε ένα URL και να τα πετάει σε κάθε refresh.
    #   το Refresh μπορεί να γίνεται με cls
    print("Version: " + version)


try:
    from bs4 import BeautifulSoup as soup
    from datetime import date
    from datetime import datetime
    from time import time, sleep
    from multiprocessing import Process, freeze_support, current_process
    from threading import Thread, current_thread
    import requests
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import os
except KeyboardInterrupt:
    sys.exit(1)
except Exception as exc:
    import sys
    print("Κάτι πάθαμε κατά το import.")
    print(str(exc))
    sys.exit(0)

""" NEW START TIME CODE """


def get_start_time():
    global start_time, start_date, start
    start = datetime.now()
    start_date = start.strftime("%d-%m-%Y")
    start_time = start.strftime("%H:%M:%S")
    print("Εκκίνηση: " + start_date + ", " + start_time)
    print("")


def get_title_time():
    elapsed_time = datetime.now() - start  # διαφορά χρόνου σε σχέση με την αρχή
    total_seconds = elapsed_time.seconds  # σύνολο διαφοράς σε δευτερόλεπτα
    total_minutes = int(total_seconds / 60)  # σύνολο διαφοράς σε λεπτά
    # υπόλοιπα δευτερόλετπα μετά τον υπολογισμό των λεπτών
    seconds = int(total_seconds - (total_minutes * 60))
    # υπόλοιπες ώρες μετά τον υπολογισμό των λεπτών
    hours = int(total_minutes / 60)
    # υπόλοιπα λετπά μετά τον υπολογισμό των ωρών
    minutes = total_minutes - (hours * 60)
    days = elapsed_time.days  # μέρες διαφοράς

    if len(str(hours)) == 1:
        hours = "0" + str(hours)
    if len(str(minutes)) == 1:
        minutes = "0" + str(minutes)
    if len(str(seconds)) == 1:
        seconds = "0" + str(seconds)
    formatted_time = str(hours) + ":" + str(minutes) + ":" + str(seconds)
    return(formatted_time)


""" NEW ELAPSED TIME CODE """


def get_elapsed_time(e):
    elapsed_time = datetime.now() - start  # διαφορά χρόνου σε σχέση με την αρχή
    total_seconds = elapsed_time.seconds  # σύνολο διαφοράς σε δευτερόλεπτα
    total_minutes = int(total_seconds / 60)  # σύνολο διαφοράς σε λεπτά
    # υπόλοιπα δευτερόλετπα μετά τον υπολογισμό των λεπτών
    seconds = int(total_seconds - (total_minutes * 60))
    # υπόλοιπες ώρες μετά τον υπολογισμό των λεπτών
    hours = int(total_minutes / 60)
    # υπόλοιπα λετπά μετά τον υπολογισμό των ωρών
    minutes = total_minutes - (hours * 60)
    days = elapsed_time.days  # μέρες διαφοράς

    if minutes == 0 and seconds == 0:
        print("Όσο πάει χειροτερεύει. Τελείωσε σε χρόνο 0")
    elif hours > 0:
        print("Όσο πάει χειροτερεύει. Τελείωσε σε " + str(hours) + " ώρες, " + str(minutes) +
              " λεπτά και " + str(seconds) + " δευτερόλεπτα (" + str(round(total_seconds, 2)) + " δευτερόλεπτα).")
    else:
        print("Όσο πάει χειροτερεύει. Τελείωσε σε " + str(minutes) + " λεπτά και " +
              str(seconds) + " δευτερόλεπτα (" + str(round(total_seconds, 2)) + " δευτερόλεπτα).")
    print("")

    if len(str(hours)) == 1:
        hours = "0" + str(hours)
    if len(str(minutes)) == 1:
        minutes = "0" + str(minutes)
    if len(str(seconds)) == 1:
        seconds = "0" + str(seconds)
    formatted_time = str(hours) + ":" + str(minutes) + \
        ":" + str(seconds) + " (H:M:S)"
    # print(formatted_time)
    ws_write.cell(row=1, column=6, value=formatted_time)


def set_files():
    global write_path, write_file, alt_write_file, wb_write, ws_write
    if os.path.exists("C:\\Users\\manager\\Desktop") == True:
        write_path = ("C:\\Users\\manager\Desktop")
    elif os.path.exists("K:\\SALES\\Stock\\Scripts\\translated") == True:
        write_path = ("K:\\SALES\\Stock\\Scripts\\translated")
    elif os.path.exists("Z:\OneDrive\eShop Stuff\PRODUCT\Product") == True:
        write_path = ("Z:\OneDrive\eShop Stuff\PRODUCT\Product")
    write_file = ('Translate_Needed_' + start_date + '.xls')
    alt_write_file = ('Translate_Needed_alt_' + start_date + '.xls')
    os.chdir(write_path)
    # wb_write = xlwt.Workbook()
    # ws_write = wb_write.add_sheet("en_exist", cell_overwrite_ok = True)
    # ws_write.write(0, 0, "ΚΩΔΙΚΟΣ")
    # ws_write.write(0, 1, "ΤΙΤΛΟΣ")
    # ws_write.write(0, 2, "GR")
    # ws_write.write(0, 3, "EN")
    # ws_write.write(0, 4, "CRAZY")
    wb_write = Workbook()
    ws_write = wb_write.active
    ws_write.title = "en_exist"
    ws_write.cell(row=1, column=1, value="ΚΩΔΙΚΟΣ")
    ws_write.cell(row=1, column=2, value="ΤΙΤΛΟΣ")
    ws_write.cell(row=1, column=3, value="GR")
    ws_write.cell(row=1, column=4, value="EN")
    ws_write.cell(row=1, column=5, value="CRAZY")
    ws_write.column_dimensions['A'].width = 12
    ws_write.column_dimensions['B'].width = 24
    ws_write.column_dimensions['C'].width = 8
    ws_write.column_dimensions['D'].width = 8
    ws_write.column_dimensions['E'].width = 24


def list_pages():
    global pages_list
    pages_list = []

    if os.path.exists("K:\\SALES\\Stock\\Scripts\\urlcheck.txt") == True:
        print("Το αρχείο urlcheck.txt βρέθηκε. Φορτώνω από εκεί.")
        text_file = open("K:\\SALES\\Stock\\Scripts\\urlcheck.txt", "r")
        lines = text_file.readlines()
        for line in lines:
            if line.find("http") == 0:
                pages_list.append(line.strip())
                print("Πρόσθεσα: " + line.strip())
        text_file.close()
    else:
        print("Δεν βρέθηκε το αρχείο.")
        page_url = input("Δώσε πράμα: ")
        if page_url == "":
            print("Εκκένωσης...")
            sys.exit(1)
        while page_url.find("http") >= 0:
            pages_list.append(page_url)
            page_url = input(
                "Έ έτσι ξεροσφύρι θα τη βγάλουμε; Δώσε κι άλλο πράμα: ")
        else:
            print("Τα μαζεύω και φεύγω.")
            print("")


def get_cy_mainpage(page_url):
    global total_next_pages, cat_page, query_mark, categories, cat_offset_url, crazy_page_soup
    print("Φανταστική σελιδούλα με όνομα:")
    print(page_url)
    start_page_soup = load_soup(page_url, wait, retries)
    # result = requests.get(page_url, cookies = cookies, headers = headers)
    # webpage = result.content
    # start_page_soup = soup(webpage, "html5lib")
    if crazy_mark == False:
        next_pages_category = start_page_soup.findAll(
            'a', {'class': 'mobile_list_navigation_link'})
        try:
            total_next_pages = next_pages_category[len(
                next_pages_category)-1].text
        except:
            total_next_pages = "1"
        print("Σύνολο σελίδων: " + str(total_next_pages))
        cat_page, query_mark, categories = str(page_url).partition("?")
        cat_offset_url = cat_page + query_mark + \
            "offset=" + str(offset) + "&" + categories
    else:
        crazy_page_soup = start_page_soup.findAll(
            'table', {'class': 'crazy-container'})
        total_next_pages = "1"
    # print("")
    # print("Offset page: " + cat_offset_url)


def get_total_products():
    global total_prod, tp
    if crazy_mark == False:
        last_offset = (int(total_next_pages) - 1) * 10
        last_cat = cat_page + query_mark + "offset=" + \
            str(last_offset) + "&" + categories
        # result = requests.get(last_cat, cookies = cookies, headers = headers)
        # webpage = result.content
        # last_page_soup = soup(webpage, "html5lib")
        last_page_soup = load_soup(last_cat, wait, retries)
        last_prod = last_page_soup.findAll(
            'table', {'class': 'web-product-container'})
        total_prod = len(last_prod) + last_offset
    else:
        # result = requests.get(page, cookies = cookies, headers = headers)
        # webpage = result.content
        # crazy_page_soup = soup(webpage, "html5lib")
        crazy_page_soup = load_soup(page, wait, retries)
        last_prod = crazy_page_soup.findAll(
            'table', {'class': 'crazy-container'})
        total_prod = len(last_prod)
    tp = total_prod
    print("Βρήκα " + str(total_prod) + " προϊόντα. Τα κεφάλια μέσα.")
    print("")


def load_soup_en(page):
    global page_soup_en
    result = requests.get(page, cookies=cookies, headers=headers)
    webpage = result.content
    page_soup_en = soup(webpage, "html5lib")
    # return(result)


def load_soup_el(page):
    global page_soup_el
    result = requests.get(page, headers=headers)
    webpage = result.content
    page_soup_el = soup(webpage, "html5lib")
    # return(result)


def multi_load_soup(page, wait, retries, lang):
    attempt = 0
    while attempt < retries:
        try:
            thread1 = Thread(target = load_soup_en, args =(page,))
            thread2 = Thread(target = load_soup_el, args =(page,))
            # proc1 = Process(target=load_soup_en(page))
            # proc2 = Process(target=load_soup_el(page))
            if lang == "EL":
                # print("proc1 starting")
                thread1.start()
                ### proc1.start()
                # print("proc2 starting")
                thread2.start()
                ### proc2.start()
                thread1.join()
                ### proc1.join()
                # print("proc1 done")
                thread2.join()
                ### proc2.join()
                # print("proc2 done")
            else:
                # print("proc1 starting")
                thread1.start()
                thread1.join()
                ### proc1.start()
                ### proc1.join()
                # print("proc1 done")
            break
        except Exception as exc:
            print("")
            print("Στο φόρτωμα της σελίδας, πέσαμε πάνω στο:")
            print(str(exc))
            print("Ξαναπροσπαθώ σε " + str(retries) + ".")
            sleep(wait)
            attempt += 1
    if attempt == retries:
        print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
        input()
        sys.exit(0)

    # if selection == "EL":
        # return(page_soup_el, page_soup_en)
    # else:
        # return(page_soup_en)
    # return(page_soup)


def load_soup(page, wait, retries):
    attempt = 0
    while attempt < retries:
        try:
            result = requests.get(page, cookies=cookies, headers=headers)
            webpage = result.content
            page_soup = soup(webpage, "html5lib")
            break
        except Exception as exc:
            print("")
            print("Στο φόρτωμα της σελίδας, πέσαμε πάνω στο:")
            print(str(exc))
            print("Ξαναπροσπαθώ σε " + str(retries) + ".")
            sleep(wait)
            attempt += 1
    if attempt == retries:
        print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
        input()
        sys.exit(0)

    return(page_soup)


def get_cy_details(container, found):
    global cy_code, en_title, en_desc_text, translated, gr_desc, prod_page_soup_el
    translated = False
    gr_desc = False
    if crazy_mark == True:
        cy_code = container.find(
            'tr', {'class': 'crazy-title-row'}).span.text.strip()
    else:
        cy_code = container.font.text.replace("(", "").replace(")", "")
    a_page = "https://www.e-shop.cy/product?id=" + cy_code

    if selection == "EL":
        # print("Loading: EL")
        multi_load_soup(a_page, wait, retries, "EL")
        while True:
            try:
                prod_page_soup_el = page_soup_el
                prod_page_soup_en = page_soup_en
                break
            except UnboundLocalError:
                sleep(1)
    else:
        # print("Loading: EN")
        multi_load_soup(a_page, wait, retries, "EN")
        while True:
            try:
                prod_page_soup_en = page_soup_en
                break
            except UnboundLocalError:
                sleep(1)

    en_title = prod_page_soup_en.h1.text.strip()
    en_desc_text = ""
    en_d_soup = prod_page_soup_en.find('div', {'id': 'mobile_desc'})
    if prod_page_soup_en.find('a', {'class': 'crazy-promo'}):
        is_crazy = True
    else:
        is_crazy = False
    en_product_table_title = prod_page_soup_en.find(
        'td', {'class': 'product_table_title'})
    if en_d_soup == None or en_d_soup.text.find('Σύνολο ψήφων') > 0 or en_product_table_title.text.strip() != "Description":
        en_desc_text = ""
        translated = False
        # print("Το " + cy_code + " δεν έχει ελληνική ούτε αγγλική περιγραφή.")
    else:
        en_desc_text = en_d_soup.decode_contents().strip()
        if en_desc_text.find('Product description is temporary unavailable in English') >= 0:
            translated = False
            # print("Το " + cy_code + " έχει ελληνική και δεν έχει αγγλική περιγραφή.")
            # print("e = " + str(e))
        else:
            translated = True
            # print("Το " + cy_code + " έχει αγγλική περιγραφή.")

    gr_desc_text = ""
    if selection == "":
        gr_desc = True
        prod_page_soup_el = ""
    elif selection == "EL":
        # prod_page_soup_el = load_soup(a_page, wait, retries, "EL")
        gr_d_soup = prod_page_soup_el.find('div', {'id': 'mobile_desc'})
        gr_product_table_title = prod_page_soup_el.find(
            'td', {'class': 'product_table_title'})
        if gr_d_soup == None or gr_d_soup.text.find('Σύνολο ψήφων') > 0 or gr_product_table_title.text.strip() != "Περιγραφή":
            gr_desc_text = ""
            gr_desc = False
        else:
            gr_desc = True

    if gr_desc == False and translated == False:
        print("Το " + cy_code + " δεν έχει ούτε ελληνική ούτε αγγλική περιγραφή.")
    elif gr_desc == True and translated == True:
        print("Το " + cy_code + " έχει ελληνική και αγγλική περιγραφή.")
    elif gr_desc == True and translated == False:
        print("Το " + cy_code + " έχει ελληνική και δεν έχει αγγλική περιγραφή.")
        found += 1
    elif gr_desc == False and translated == True:
        print("Το " + cy_code + " δεν έχει ελληνική και έχει αγγλική περιγραφή.")
        found += 1

    # print("")
    return(is_crazy, found)


def write_results(e):
    font_green = Font(color="8FCE00")
    font_red = Font(color="F44336")

    ws_write.cell(row=e, column=1, value=cy_code)
    ws_write.cell(row=e, column=2, value=en_title)
    if selection == "EL":
        if gr_desc == False:
            gr_desc_value = "NO GR"
            ws_write.cell(row=e, column=3).font = font_red
        else:
            gr_desc_value = "OK"
            ws_write.cell(row=e, column=3).font = font_green
    else:
        gr_desc_value = "-"
    ws_write.cell(row=e, column=3, value=gr_desc_value)
    if translated == False:
        translate_value = "NO EN"
        ws_write.cell(row=e, column=4).font = font_red
    else:
        translate_value = "OK"
        ws_write.cell(row=e, column=4).font = font_green
    ws_write.cell(row=e, column=4, value=translate_value)
    if page.find("crazysundays") >= 0 or crazy_mark == True or is_crazy == True:
        crazy_value = "CRAZY"
    else:
        crazy_value = "-"
    ws_write.cell(row=e, column=5, value=crazy_value)
    if warranty_text != "":
        ws_write.cell(row=e, column=6, value=warranty_text)
    else:
        ws_write.cell(row=e, column=6, value="-")
    # print("")


def write_it_down():
    try:
        wb_write.save(write_file)
        print(write_file + ", το έχω γραμμένο στο " + write_path)
    except:
        wb_write.save(alt_write_file)
        print("Πιθανώς κάποιος παίζει με το αρχείο. Προχωράω στο παρασύνθημα.")
        print(alt_write_file + ", το έχω γραμμένο στο " + write_path)


def initialize():
    answer = "YES"
    cookies = {'language': 'en'}
    headers = {
        'User-Agent': "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.27 Safari/537.17"}
    wait = 5
    retries = 3
    attempt = 0
    crazy_mark = False
    is_crazy = False
    e = 2
    found = 0
    warranty_text = ""
    return(answer, cookies, headers, wait, retries, attempt, crazy_mark, is_crazy, e, found, warranty_text)


def aarrg():
    war_selection = False
    selection = ""
    if len(sys.argv) > 1:
        for arg in sys.argv:
            selection = ""
            if arg.find("-el") == 0 or arg.find("-EL") == 0:
                selection = "EL"
                print("Βρήκα '" + selection + "'. Θα ψάξω και στα ελληνικά.")

            if arg.find("-war") == 0 or arg.find("-WAR") == 0:
                print("Βρήκα 'WAR'. Θα γίνει πόλεμος εγγύησης αλά ελληνικά.")
                war_selection = True
                selection = "EL"
    else:
        selection = ""
        war_selection = False

    return(selection, war_selection)


def get_warranty(page_soup):
    global warranty_text
    warranty_text = ""
    cy_desc_text = ""
    rest = ""
    # assign the product_table_body soup
    cy_d_soup = page_soup.find('td', {'class': 'product_table_body'})
    # assign the product_table_title soup
    cy_product_table_title = page_soup.find(
        'td', {'class': 'product_table_title'})
    # if product_table_body is empty or contains votes or product_table_title doesn't contain Περιγραφή then there is no description
    if cy_d_soup == None or cy_d_soup.text.find('Σύνολο ψήφων') > 0 or cy_product_table_title.text.strip() != "Περιγραφή":
        cy_desc_text = ""
    else:
        cy_desc_text = cy_d_soup.decode_contents().strip().replace('\n', '').replace('\t', '').replace(
            "<br/>", "<br>").replace(".gr", "")  # decode description content replace wrong html values and any .gr mentions
        if cy_desc_text.find('Eγγύηση') >= 0:
            string, warranty, rest = cy_desc_text.rpartition('Eγγύηση:')
        elif cy_desc_text.find('Εγγύηση') >= 0:
            string, warranty, rest = cy_desc_text.rpartition('Εγγύηση:')
        else:
            rest = ""
        warranty_text = rest.strip().replace(
            '<a class="navy_link" href="https://www.e-shop.cy/support#doa">DOA 7 ημερών</a>', '')


if __name__ == '__main__':
    freeze_support()
    try:
        get_start_time()
        answer, cookies, headers, wait, retries, attempt, crazy_mark, is_crazy, e, found, warranty_text = initialize()
        print("Επιλογές:")
        selection, war_selection = aarrg()
        # print("selection: " + selection)
        print("Done.")
        print("")
        set_files()
        os.system("title " + "Μαζεύω σελίδες...")
        list_pages()
        ### for page in pages_list:
        for p in range(0, len(pages_list)):
            page = pages_list[p]
            cur_page = p + 1
            offset = 0  # starting offset value set to 0 and in each for loop, 10 will be added
            if page.find("crazysundays") >= 0:
                crazy_mark = True
            else:
                crazy_mark = False
            get_cy_mainpage(page)
            print("")
            os.system("title " + "Μαζεύω προϊόντα...")
            get_total_products()
            for q in range(0, int(total_next_pages)):
                single_page_soup = load_soup(cat_offset_url, wait, retries)
                if crazy_mark == False:
                    containers = single_page_soup.findAll(
                        'table', {'class': 'web-product-container'})
                else:
                    containers = single_page_soup.findAll(
                        'table', {'class': 'crazy-container'})

                for container in containers:
                    attempt = 0
                    tp = tp - 1
                    current_prod = total_prod - tp
                    rem_prod = total_prod - current_prod
                    title_time = get_title_time()
                    print_text = "Σελίδα: " + str(cur_page) + "/" + str(len(pages_list)) + ". Προϊόν: " + str(
                        current_prod) + "/" + str(total_prod) + ". " + title_time
                    os.system("title " + print_text)
                    if total_prod - (total_prod - tp) > 0:
                        print("Τα πίνω με το: " + str(current_prod) + "/" +
                              str(total_prod) + ". Έχω ακόμα: " + str(rem_prod))
                    else:
                        print("Τα πίνω με το: " + str(current_prod) +
                              "/" + str(total_prod) + ".")
                    while attempt < 3:
                        try:
                            # e = get_cy_details(container, e)
                            is_crazy, found = get_cy_details(container, found)
                            if prod_page_soup_el != "":
                                get_warranty(prod_page_soup_el)
                            else:
                                pass
                            # print("Translated: " + str(translated))
                            # print("gr_desc: " + str(gr_desc))
                            if war_selection == True:
                                write_results(e)
                                e += 1
                            elif translated == False or gr_desc == False:
                                write_results(e)
                                e += 1
                            print("")
                            break
                        except Exception as exc:
                            print("")
                            print("Όχι ρε φίλε. Μόλις σκόνταψα γιατί:")
                            print(str(exc))
                            print(
                                "Κάτσε να σκουπιστώ και ξαναπροσπαθώ σε 3 δεύτερα.")
                            attempt += 1
                            sleep(5)
                        if attempt >= 3:
                            print("")
                            print(
                                "Ρε φίλε προσπάθησα 3 φορές. Φαίνεται δεν ταιριάζουμε. Να περάσει ο επόμενος.")
                            print("")
                            continue
                if crazy_mark == True:
                    continue
                else:
                    offset += 10
                    cat_offset_url = cat_page + query_mark + \
                        "offset=" + str(offset) + "&" + categories

        get_elapsed_time(e)
        # print("e: " + str(e))
        if e > 2:
            write_it_down()
        else:
            pass
        if found > 1:
            input("Βρήκα " + str(found) +
                  " να κοιμούνται. Παίξε λίγο με το πράμα σου για να κλείσω.")
        elif found == 1:
            input("Βρήκα " + str(found) +
                  " να κοιμάται. Παίξε λίγο με το πράμα σου για να κλείσω.")
        elif found == 0:
            input(
                "Είσαι 'νταξ'. Δεν κοιμάται κανένας τους. Παίξε λίγο με το πράμα σου για να κλείσω.")
        print("")
    except KeyboardInterrupt as exc:
        # os.system('cls')
        print("")
        print("Ρε μην τον παίζεις έχουμε δουλειά!")
        input("Τι να σε κάνω;")
        print("")
        sys.exit(0)
    except Exception as exc:
        print(str(exc))
        input("Κάτι δεν πάει καλά. Δες πιο πάνω τι συνέβη και ξανατρέξε.")
        sys.exit(0)
