import sys


def ti_paizei():
    version = "Version 1.1 beta"
    # gr_cy_mobiles.py
    # Ψάχνει για όλα τα κινητά στο ελληνικό site
    # και αν υπάρχει στο κυπριακό ελέγχει αν υπάρχουν specs
    # Current Version 1.1 beta
    #########
    # Changelog Version 1.1 beta
    # - Ανασχεδιασμός κώδικα σε διάφορα σημεία
    # - Χρήση openpyxl
    # - Προσπάθεια να βρίσκει τις διαφορές στα προϊόντα
    #########
    # ToDo
    #########
    """ κολλάει σε ένα συγκεκριμένο προϊόν: TEL.092759 """
    """ φτιάχτηκε αλλά κολλάει σε ένα exception """
    """ φτιάχτηκε  """
    # Να βρίσκει μόνο τις διαφορές από το GR σε σχέση με το CY
    """ δουλεύει μόνο στα προϊόντα χωρίς περιγραφή ελέγχει μόνο yes/no  """
    # Να κρατάει το HTML format στα extra και το <br>
    # Μέγιστος αριθμός χαρακτήρων στο CPU και CAMERA = 50
    # Ανάλυση camera και αφαίρεση ότι δεν χρειάζεται
    """ έγινε προσπάθεια """
    """ μεγαλύτερο από 50 στο Οπίσθια: 12MP (f1.5-2.4/Dual Pixel PDAF, OIS) + 12MP (f2.4) + 16MP (f2.2). Εμπρόσθια: 10MP (f1.9/D PDAF) """
    """ δεν λειτουργεί στο Οπίσθια: 13MP /1.4''/autofocus/LED flash/panorama/HDR. Εμπρόσθια: 8MP /1.4'' """
    """ δεν λειτουργεί στο Οπίσθια: 16MP /f1.7/PDAF/LED flash/panorama/HDR. """
    """ δεν λειτουργεί στο Οπίσθια: 16MP (f1.7) + 5MP (f1.9), phase detect autofocus/LED flash. Εμπρόσθια: 24MP /f1.9 """
    """ δεν λειτουργεί στο Οπίσθια: Dual: 12MP (f1.8, OIS) + 20MP (f1.6) phase & laser autofocus/dual-LED. Εμπρόσθια: 24MP """
    """ δεν λειτουργεί στο Οπίσθια: 12.2MP(f1.7/PDAF/OIS/LED flash/Auto-HDR. Εμπρόσθια: 8MP(f2.0) """
    """ δεν λειτουργεί στο Οπίσθια: 40MP(f1.6) + 20MP(f2.2/16mm) + 8MP (f3.4/80mm) + TOF camera. Εμπρόσθια: 32MP (f2.0) """
    """ δεν λειτουργεί στο Οπίσθια: 40MP(f1.6) + 8MP(f2.4) + 40MP(f1.8) + TOF 3D, Leica optics, dual-LED. Εμπρόσθια: 32MP(f2.0) """
    """ δεν λειτουργεί Mpixels στο 12MP (f1.8/26mm/OIS/PDAF) + 12MP (f2.4/52mm/OIS/PDAF/2x opt zoom),Quad-LED. Εμπρόσθια: 7MP (f2.2/32mm) """
    """ δεν λειτουργεί στο Οπίσθια: 16MP (f1.7) + 5MP (f1.9), phase detect autofocus/LED flash. Εμπρόσθια: 24MP /f1.9 """
    """ δεν λειτουργεί στο Οπίσθια: 13MP /PDAF/LED flash/panorama/HDR. Εμπρόσθια: 5MP /f2.2 """
    """ δεν λειτουργεί στο Οπίσθια: 13MP /PDAF/LED flash/panorama/HDR. Εμπρόσθια: 5MP /f2.2 """
    """ δεν λειτουργεί στο Οπίσθια: 32MP (PDAF) + 5 MP F2.2 + 2 MP F2.4 + 2 MP F2.4, Μπροστα 32MP """
    """ δεν λειτουργεί στο Οπίσθια: 13MP /f1.8/PDAF/LED flash/panorama/HDR. Εμπρόσθια: 5MP /f2.2 """
    """ δεν λειτουργεί στο Οπίσθια: 13MP /f2.2/0.5A LED flash. Εμπρόσθια: 8MP """
    """ δεν λειτουργεί το mpixels στο : Οπίσθια:13MP(f2.0/PDAF) + 5MP(f2.2) + 2MP(f2.4) + 2MP(f2.4), LED flash, HDR. Εμπρόσθια: 8MP """
    """ δεν λειτουργεί το mpixels στο : Οπίσθια:108MP(f1.7/PDAF/OIS) + 13MP(f2.4) + 5MP(f2.4), LED flash, HDR. Εμπρόσθια: 20MP(f2.2) """
    """ δεν λειτουργεί στο Οπίσθια: 48MP(f1.8/PDAF) + 8MP(f2.3) + 2MP(f2.4) LED flash, HDR. Εμπρόσθια: 8MP(f2.0) """
    # Ανάλυση CPU και αφαίρεση ότι δεν χρειάζεται
    """ έγινε προσπάθεια - μέγιστος αριθμός χαρακτήρων 50 ?"""
    """ δεν λειτουργεί στο Octa-core (4x2.0 GHz Cortex-A55 & 4x2.0 GHz Cortex-A55), GPU:Mali-G52 """
    """ δεν λειτουργεί στο Octa-core (2 x 2.7GHz & 2 x 2.3GHz & 4 x 1.9GHz), GPU:Mali-G76 MP12 """
    """ δεν λειτουργεί στο Hexa-core 2x Vortex + 4x Tempest, GPU:Apple GPU 4-core graphics """
    """ δεν λειτουργεί στο Octa-core (2x2.73 GHz & 2x2.4 GHz & 4x1.9 GHz), GPU:Mali-G76 MP12 """
    # Αφαίρεση εγγύησης από τα extra
    """ έγινε προσπάθεια - φαίνεται πως δουλεύει """
    """ δουλεύει τις περισσότερες φορές"""
    # Αφαίρεση περιττών κενών στα extra
    """ έγινε προσπάθεια  - φαίνεται πως δουλεύει (replace("\n")) """
    # Ξεχωριστά boolean ή ναι/όχι στη λίστα για τα check boxes
    """ έγινε προσπάθεια - φαίνεται πως δουλεύει (ναι/όχι) """
    # Να γράφει το πεδίο στη λίστα ακόμα και αν είναι κενά τα specs του
    """ έγινε προσπάθεια - δουλεύει """
    # Λίστα προτεραιότητας specs:
    """ έγινε προσπάθεια - φαίνεται πως δουλεύει 
    # Διαστάσεις
    # Βάρος
    # Χρόνος Ομιλίας
    # Xρόνος Αναμονής
    # Οθόνη
    # Κάρτα μνήμης
    # Mobile Internet
    # Ασύρματη επικοινωνία
    # Camera
    # Ειδοποιήσεις
    # MPixels
    # Εσωτερική μνήμη
    # Video
    # Μνήμη RAM
    # Μπαταρία	Τύπος
    # Ημερομηνία κυκλοφορίας	ΜΗΝΑΣ	ΧΡΟΝΙΑ
    # Extra
    # MMC
    # JAVA
    # NFC
    # GPS
    # Radio
    # Fingerprint
    # Αποσπώμενη μπαταρία
    # CPU
    # Ενσωματωμένοι Αισθητήρες
    # Λειτουργικό Σύστημα
    # ΠΕΡΙΓΡΑΦΗ
    """
    print(version)


try:
    from bs4 import BeautifulSoup as soup
    from random import randint
    from time import sleep as nani
    from datetime import datetime
    import requests
    import os
    import sys
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    import xlwt  # , unicodedata
except KeyboardInterrupt:
    sys.exit(0)
except Exception as exc:
    print("Κάτι πάθαμε κατά το import.")
    print(str(exc))
    sys.exit(0)


def files_setup(today_format):
    global wb_write, ws_write, write_file, alt_write_file, write_path
    try:
        if os.path.exists('Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product') == True:
            write_path = ('Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product')
        elif os.path.exists('Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product') == True:
            write_path = ('Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product')
        os.chdir(write_path)
        write_file = today_format + "_GRvsCY_mobiles_results.xls"  # name of xls write file
        # alternate name of xls write file
        alt_write_file = today_format + "_GRvsCY_mobiles_results_alt.xls"
        print("Το αρχείο θα είναι: " + write_file)
        print("")
        wb_write = Workbook()  # Create a virtual workbook to keep data in
        # add 1st sheet in virtual workbook
        ws_write = wb_write.active
        ws_write.title = "Y-N"
        ws_write.cell(row=1, column=1, value="ΚΩΔΙΚΟΣ")
        ws_write.cell(row=1, column=2, value="Y-N")
    except Exception as exc:
        print("Δεν κατάφερα να γράψω το αρχείο. Έχουμε δικαιώματα;")
        print(str(exc))
        print("")


def no_check():
    global no_check_list
    no_check_list = []
    txt_relative_path = "Mobiles\\no_check.txt"
    text_full_path = os.path.join(write_path, txt_relative_path)
    # print(text_file_path)
    if os.path.exists(text_full_path) == True:
        text_file = open(text_full_path, "r")
        # text_file = open(r"Z:\OneDrive\eShop Stuff\PRODUCT\Product\Mobiles\no_check.txt","r")
        lines = text_file.readlines()
        for line in lines:
            if line != "\n":
                no_check_list.append(line.strip())
                print("Δεν ελέγχω: " + line.strip())
        text_file.close()
    else:
        print("Δεν βρήκα το αρχείο εξαίρεσης. Προχωράμε.")
    print("")


def load_soup(page, wait, retries):
    # temp_product = page[page.rfind("=") + 1:]
    # print("Loading soup for " + temp_product)
    # print("")
    # print("Μέσα στη σούπα.")
    attempt = 0
    while attempt < retries:
        try:
            result = requests.get(page, headers=headers)
            webpage = result.content
            page_soup = soup(webpage, "html5lib")
            break
            # print("Έξω από τη σούπα.")
            # print("")
        except Exception as exc:
            print("")
            print("Στο φόρτωμα της σελίδας, πέσαμε πάνω στο:")
            print(str(exc))
            print("Ξαναπροσπαθώ σε " + str(retries) + ".")
            nani(wait)
            attempt += 1
    if attempt == retries:
        print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
        input()
        sys.exit(0)

    return(page_soup)


def get_all_products(page_url, page_soup, attempt, retries):
    while attempt < retries:
        try:
            offset = 0
            cat_pages = []
            prod_count = page_soup.find(
                'div', {'class': 'web-product-num'}).text
            prod_count = int(prod_count[:prod_count.find(" ")].strip())
            total_next_pages = int(prod_count / 10) + 1
            cat_page, query_mark, categories = str(page_url).partition("?")
            while offset < prod_count:
                # while offset < 3 :
                # print("inside while loop")
                cat_pages.append(cat_page + query_mark +
                                 "offset=" + str(offset) + "&" + categories)
                offset += 10
                # print(str(offset))
            print("Συνολο σελιδών/cat_pages: " +
                  str(total_next_pages) + "/" + str(len(cat_pages)))
            # print("Σύνολο cat_pages: " + str(len(cat_pages)))
            # p = 0
            cat_attempt = 0
            for p, page in enumerate(cat_pages):
                # for idx in range(2, 4) :
                # page = cat_pages[idx]
                try:
                    # p += 1
                    print_text = "Μετρώντας τα προϊόντα της σελίδας: " + str(p)
                    os.system("title " + "Getting page " + str(p) +
                              "/" + str(len(cat_pages)) + " items")
                    if p != len(cat_pages):
                        print(print_text, end='\r')
                    else:
                        print(print_text)
                    single_page_soup = load_soup(page, wait, retries)
                    containers = single_page_soup.findAll(
                        'table', {'class': 'web-product-container'})
                    for container in containers:
                        gr_code = container.font.text.replace(
                            "(", "").replace(")", "")
                        all_products.append(gr_code)
                except Exception as exc:
                    cat_attempt += 1
                    print("")
                    print("Ώπα πέσαμε πάνω στο:")
                    print(str(exc))
                    print("Ξαναπροσπαθώ σε 3.")
                    nani(wait)
                if cat_attempt == retries:
                    break
            break
        except Exception as exc:
            attempt += 1
            print("")
            print("Ώπα πέσαμε πάνω στο:")
            print(str(exc))
            print("Ξαναπροσπαθώ σε 3.")
            nani(wait)
    if attempt == retries:
        print("Προσπάθησα " + str(retries) + " φορές και δεν τα κατάφερα.")
        input()
        sys.exit(0)


def write_it_down(e, null):
    # print("Γράφω: " + str(e))
    if null == 0:
        try:
            wb_write.save(write_file)
        except:
            wb_write.save(alt_write_file)
    elif e > 1 or null != 0:
        try:
            wb_write.save(write_file)
            print(write_file + ", το έχω γραμμένο στο " + write_path)
        except:
            print("Πιθανώς κάποιος παίζει με το αρχείο. Προχωράω στο παρασύνθημα.")
            wb_write.save(alt_write_file)
            print(alt_write_file + ", το έχω γραμμένο στο " + write_path)
    else:
        print("Δεν έχει γίνει καμία αλλαγή στο αρχείο.")


def get_specs(gr_soup, cy_soup):
    """ Find all specs in code, the else part is for the old html code"""
    gr_specs = []
    gr_specs1 = []
    gr_specs2 = []
    cy_specs = []
    cy_specs1 = []
    cy_specs2 = []

    if gr_soup.find('td', {'class': 'product_table_body'}).findAll('td', {'class': 'details2'}):
        gr_specs1 = gr_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details1'})
        gr_specs2 = gr_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details2'})
    else:
        gr_specs = gr_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details1'})
        for i in range(0, len(gr_specs), 2):
            gr_specs2.append(gr_specs[i])
        for i in range(1, len(gr_specs), 2):
            gr_specs1.append(gr_specs[i])

    if cy_soup.find('td', {'class': 'product_table_body'}).findAll('td', {'class': 'details2'}):
        cy_specs1 = cy_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details1'})
        cy_specs2 = cy_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details2'})
    else:
        cy_specs = cy_soup.find('td', {'class': 'product_table_body'}).findAll(
            'td', {'class': 'details1'})
        for i in range(0, len(cy_specs), 2):
            cy_specs2.append(cy_specs[i])
        for i in range(1, len(cy_specs), 2):
            cy_specs1.append(cy_specs[i])

    if len(cy_specs1) == len(gr_specs1):
        print("len(cy_specs1 / gr_specs1): " + str(len(cy_specs1)))
    else:
        print("len(cy_specs1): " + str(len(cy_specs1)))
        print("len(gr_specs1): " + str(len(gr_specs1)))

    if len(cy_specs2) == len(gr_specs2):
        print("len(cy_specs2 / gr_specs2): " + str(len(cy_specs2)))
    else:
        print("len(cy_specs2): " + str(len(cy_specs2)))
        print("len(gr_specs2): " + str(len(gr_specs2)))

    return(gr_specs1, gr_specs2, cy_specs1, cy_specs2)


def compare(gr_specs1, gr_specs2, cy_specs1, cy_specs2):
    txt_gr_specs1 = []
    txt_gr_specs2 = []
    txt_cy_specs1 = []
    txt_cy_specs2 = []
    
    for spec in gr_specs1:
        try:
            print(spec)
            txt_gr_specs1.append(spec.text.strip())
        except Exception as exc:
            txt_gr_specs1.append(spec)
    
    for spec in txt_gr_specs1:
        print(spec)

    for spec in gr_specs2:
        try:
            txt_gr_specs2.append(spec.text.strip())
        except Exception as exc:
            txt_gr_specs2.append(spec)

    for spec in cy_specs1:
        try:
            txt_cy_specs1.append(spec.text.strip())
        except Exception as exc:
            txt_cy_specs1.append(spec)

    for spec in cy_specs2:
        try:
            txt_cy_specs2.append(spec.text.strip())
        except Exception as exc:
            txt_cy_specs2.append(spec)

    dif_found = 0
    for idx in range(len(cy_specs1)):
        print("cy_specs1: " + cy_specs1[idx], "gr_specs1: " + gr_specs1[idx])
        if cy_specs1[idx] != gr_specs1[idx]:
            print("Βρήκα διαφορά στο " + str(gr_specs2[idx]))
            dif_found += 1
            if dif_found == 1:
                ws_write_product = wb_write.create_sheet(product)
                ws_write_product.cell(row=1, column=1, value="ΤΙΤΛΟΣ")
                ws_write_product.cell(row=1, column=2, value="ΠΕΡΙΓΡΑΦΗ")
            ws_write_product.cell(row=dif_found+1, column=1, value=gr_specs1[idx])
            ws_write_product.cell(row=dif_found+1, column=2, value=gr_specs2[idx])
        else:
            pass
            # print("Same")


def get_description(page_soup):
    desc_text = ""
    desc_soup = page_soup.find('td', {'class': 'product_table_body'})
    product_table_title = page_soup.find(
        'td', {'class': 'product_table_title'})
    if desc_soup == None or desc_soup.text.find('Σύνολο ψήφων') > 0 or product_table_title.text.strip() != "Περιγραφή":
        desc_text = ""
    else:
        desc_text = desc_soup.decode_contents().strip().replace('\n', '').replace(
            '\t', '').replace("<br/>", "<br>").replace(".gr", "")
        temp_text = desc_text.partition(
            '<table border="0" cellpadding="0" cellspacing="0"')[0]
        if temp_text == "<br>":
            desc_text = ""
        else:
            desc_text = temp_text
    return(desc_text)


def add_specs(gr_specs1, gr_specs2):
    for i in range(len(gr_specs1)):
        # print(gr_specs1[i].text)
        # if gr_specs1[i].text.strip().find("2 χρόνια ") >= 0 gr_specs1[i].text.strip().find("1 χρόνος ") >= 0 :
        temp_title = gr_specs2[i].text.strip()
        temp_spec = gr_specs1[i].text.strip()

        if temp_title.find("Διαστάσεις") >= 0:
            all_specs["01. Διαστάσεις"] = temp_spec.replace("\n", "")
        if temp_spec.find("2 χρόνια ") >= 0 or temp_spec.find("1 χρόνος ") >= 0:
            print("Warranty found")
            # temp_spec = gr_specs1[i].text.strip()
            if temp_spec.find("2 χρόνια ") >= 0:
                extra = temp_spec.rpartition("2 χρόνια ")[0]
            elif temp_spec.find("1 χρόνος ") >= 0:
                extra = temp_spec.rpartition("1 χρόνος ")[0]
            spec_specs.append(extra.strip())
        else:
            spec_specs.append(temp_spec.replace("\n", ""))
            # spec_specs.append(gr_specs1[i].text.strip().replace("\n", ""))
        # print(gr_specs1[i].text.strip().replace("\n", ""))
    return(spec_specs)


def merge_specs(spec_title, spec_specs):
    for i in range(len(spec_title)):
        # print("i: " + str(i))
        # print(spec_title[i] + ": " + spec_specs[i])
        if spec_specs[i].find("mAh") >= 0:
            print("found mAh")
            battery_cap, delim, rest = spec_specs[i].partition("mAh")
            all_specs["15. Μπαταρία"] = battery_cap.strip()
            if rest.strip().find("Μη αποσπώμενη") == 0:
                all_specs["26. Removable"] = "Οχι"
            elif rest.strip().find("Αποσπώμενη") == 0:
                all_specs["26. Removable"] = "Ναί"
            else:
                battery_type, delim, removable = rest.strip().partition(" ")
                all_specs["16. Τύπος"] = battery_type.strip()
                if removable.strip().find("Μη αποσπώμενη") == 0:
                    all_specs["26. Removable"] = "Οχι"
                elif removable.strip().find("Αποσπώμενη") == 0:
                    all_specs["26. Removable"] = "Ναί"
        elif spec_title[i].find("Ημερομηνία κυκλοφορίας") >= 0:
            month, partition, year = spec_specs[i].partition("-")
            all_specs["17. Μήνας"] = month
            all_specs["18. Χρόνος"] = year
        else:
            all_specs[spec_title[i]] = spec_specs[i]
    return(all_specs)


def spec_fixes(all_specs):
    if all_specs.get("20. MMC") is None:
        all_specs["20. MMC"] = "Οχι"
    if all_specs.get("21. JAVA") is None:
        all_specs["21. JAVA"] = "Οχι"
    if all_specs.get("22. NFC") is None:
        all_specs["22. NFC"] = "Οχι"
    if all_specs.get("23. GPS") is None:
        all_specs["23. GPS"] = "Οχι"
    if all_specs.get("24. Radio") is None:
        all_specs["24. Radio"] = "Οχι"
    if all_specs.get("25. Fingerprint") is None:
        all_specs["25. Fingerprint"] = "Οχι"

    if "09. Camera" in all_specs and all_specs["09. Camera"] != "Ναι":
        print("found 09. Camera")
        """mpixels fix procedure"""
        mpixels_text = all_specs["09. Camera"]
        mpixels_temp = mpixels_text[:mpixels_text.find("MP")]
        if mpixels_temp.find(" ") >= 0:
            mpixels = mpixels_temp[mpixels_temp.find(" ") + 1:]
        else:
            mpixels = mpixels_temp[:mpixels_temp.find("MP")]
        mpixels = mpixels.strip()
        prefix = "11. "
        all_specs[prefix + "MPixels"] = mpixels
        """camera fix procedure"""
        camera_text = all_specs.get("09. Camera")
        print("camera_text: " + camera_text)
        while camera_text.find("(") >= 0:
            ctext1 = camera_text[:camera_text.find("(")].strip()
            ctext2 = camera_text[camera_text.find(")") + 1:].strip()
            print("ctext1: " + ctext1)
            print("ctext2: " + ctext2)
            # υπάρχει το "(" αλλά δεν υπάρχει το ")"
            if ctext2.find("(") >= 0 and ctext2.find(")") < 0:
                # υπάρχει τελεία και τα MP είναι μακριά
                if ctext2.find(".") > 0 and ctext2.find("MP") > 15:
                    ctext2 = ctext2[ctext2.find("."):]
                else:
                    ctext2 = ""
            print("ctext2: " + ctext2)
            camera_text = ctext1.strip() + ctext2.strip()
            print("camera_text: " + camera_text)
            # input()
            print("camera_text.find(',') >= 0 = " +
                  str(camera_text.find(",") >= 0))

        while camera_text.find(",") >= 0:
            ctext1 = camera_text[:camera_text.find(",")].strip()
            ctext2 = camera_text[camera_text.find(","):]
            if ctext2.find(".") >= 0:
                ctext2 = ctext2[ctext2.find("."):].strip()
            else:
                ctext2 = ""
            print("ctext1: " + ctext1)
            print("ctext2: " + ctext2)
            camera_text = (ctext1 + ctext2).replace("+ ", " + ")
            # input()
            print("camera_text: " + camera_text)

        while camera_text.find("/") >= 0:
            ctext1 = camera_text[:camera_text.find("/")].strip()
            ctext2 = camera_text[camera_text.find("/"):].strip()
            if ctext2.find(".") >= 0:
                ctext2 = ctext2[ctext2.find("."):].strip()
            elif ctext2.find("/") >= 0:
                ctext2 = ctext2[:ctext2.find("/")].strip()
            else:
                continue
            print("ctext1: " + ctext1)
            print("ctext2: " + ctext2)
            camera_text = (ctext1 + ctext2).replace("+ ", " + ")
            # input()
            print("camera_text: " + camera_text)
        all_specs["09. Camera"] = camera_text.replace("  ", " ").strip()

    """CPU fix procedure"""
    if "27. CPU" in all_specs:
        cpu_text = all_specs.get("27. CPU")
        print("cpu_text: " + cpu_text)
        cpu_text = cpu_text.replace("(", "")
        cpu_text = cpu_text.replace(")", "")
        if len(cpu_text) >= 50:
            cpu_text = cpu_text.replace("GHz & ", "& ")
        all_specs["27. CPU"] = cpu_text
        print("cpu_text: " + cpu_text)
    """talk time fix procedure"""
    if "03. Χρόνος Ομιλίας" in all_specs:
        talk_text = all_specs.get("03. Χρόνος Ομιλίας")
        print("talk_text: " + talk_text)
        talk_text = talk_text[:talk_text.find(" ")].strip()
        all_specs["03. Χρόνος Ομιλίας"] = talk_text
        print("talk_text: " + talk_text)
    if "04. Xρόνος Αναμονής" in all_specs:
        wait_text = all_specs.get("04. Xρόνος Αναμονής")
        print("wait_text: " + wait_text)
        wait_text = wait_text[:wait_text.find(" ")].strip()
        all_specs["04. Xρόνος Αναμονής"] = wait_text
        print("wait_text: " + wait_text)
    """memory fix procedure"""
    if "12. Εσωτερική μνήμη" in all_specs:
        storage_text = all_specs.get("12. Εσωτερική μνήμη")
        print("storage_text: " + storage_text)
        storage_text = storage_text[:storage_text.find(" ")].strip()
        all_specs["12. Εσωτερική μνήμη"] = storage_text
        print("storage_text: " + storage_text)
    if "14. Μνήμη RAM" in all_specs:
        ram_text = all_specs.get("14. Μνήμη RAM")
        print("ram_text: " + ram_text)
        ram_text = ram_text[:ram_text.find(" ")].strip()
        all_specs["14. Μνήμη RAM"] = ram_text
        print("ram_text: " + ram_text)
    # """extras fix procedure"""
    # if "19. Extra" in all_specs :
        # extra = all_specs.get("19. Extra")
        # extra_text, partition, rest = extra.rpartition("<br>")
        # all_specs["19. Extra"] = extra_text
    """ adding empty fields before sorting """
    if len(all_specs) < 29:
        if all_specs.get("01. Διαστάσεις (mm)") == None:
            all_specs["01. Διαστάσεις (mm)"] = ""
        if all_specs.get("02. Βάρος (γραμ.)") == None:
            all_specs["02. Βάρος (γραμ.)"] = ""
        if all_specs.get("03. Χρόνος Ομιλίας") == None:
            all_specs["03. Χρόνος Ομιλίας"] = ""
        if all_specs.get("04. Xρόνος Αναμονής") == None:
            all_specs["04. Xρόνος Αναμονής"] = ""
        if all_specs.get("05. Οθόνη") == None:
            all_specs["05. Οθόνη"] = ""
        if all_specs.get("06. Κάρτα μνήμης") == None:
            all_specs["06. Κάρτα μνήμης"] = ""
        if all_specs.get("07. Mobile Internet") == None:
            all_specs["07. Mobile Internet"] = ""
        if all_specs.get("08. Ασύρματη επικοινωνία") == None:
            all_specs["08. Ασύρματη επικοινωνία"] = ""
        if all_specs.get("09. Camera") == None:
            all_specs["09. Camera"] = ""
        if all_specs.get("10. Ειδοποιήσεις") == None:
            all_specs["10. Ειδοποιήσεις"] = ""
        if all_specs.get("11. MPixels") == None:
            all_specs["11. MPixels"] = ""
        if all_specs.get("12. Εσωτερική μνήμη") == None:
            all_specs["12. Εσωτερική μνήμη"] = ""
        if all_specs.get("13. Video") == None:
            all_specs["13. Video"] = ""
        if all_specs.get("14. Μνήμη RAM") == None:
            all_specs["14. Μνήμη RAM"] = ""
        if all_specs.get("15. Μπαταρία") == None:
            all_specs["15. Μπαταρία"] = ""
        if all_specs.get("16. Τύπος") == None:
            all_specs["16. Τύπος"] = ""
        if all_specs.get("17. Μήνας") == None:
            all_specs["17. Μήνας"] = ""
        if all_specs.get("18. Χρόνος") == None:
            all_specs["18. Χρόνος"] = ""
        if all_specs.get("19. Extra") == None:
            all_specs["19. Extra"] = ""
        if all_specs.get("20. MMC") == None:
            all_specs["20. MMC"] = ""
        if all_specs.get("21. JAVA") == None:
            all_specs["21. JAVA"] = ""
        if all_specs.get("22. NFC") == None:
            all_specs["22. NFC"] = ""
        if all_specs.get("23. GPS") == None:
            all_specs["23. GPS"] = ""
        if all_specs.get("24. Radio") == None:
            all_specs["24. Radio"] = ""
        if all_specs.get("25. Fingerprint") == None:
            all_specs["25. Fingerprint"] = ""
        if all_specs.get("26. Removable") == None:
            all_specs["26. Removable"] = ""
        if all_specs.get("27. CPU") == None:
            all_specs["27. CPU"] = ""
        if all_specs.get("28. Ενσωματωμένοι Αισθητήρες") == None:
            all_specs["28. Ενσωματωμένοι Αισθητήρες"] = ""
        if all_specs.get("29. Λειτουργικό Σύστημα") == None:
            all_specs["29. Λειτουργικό Σύστημα"] = ""
    return(all_specs)


def sort_specs(all_specs):
    sorted_specs = {}
    for x, y in sorted(all_specs.items()):
        # print("Sorting " + x.strip() + ": " + y.strip().replace("\n", ""))
        if x.find(":") >= 0:
            sorted_specs[x[:-1]] = y.strip().replace("\n", "")
        else:
            sorted_specs[x] = y.strip().replace("\n", "")
    return(sorted_specs)


def initialize():
    start = datetime.now()
    today_format = start.strftime("%y-%m-%d")
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.27 Safari/537.17'}
    offset = 0
    attempt = 0
    retries = 5
    wait = 3
    found = 0
    yn = 0
    if len(sys.argv) > 1 and sys.argv[1] == "-saveit":
        do_i_save = True
    else:
        do_i_save = False
    print("do_i_save: " + str(do_i_save))
    total = 0
    all_products = []
    prod_count = 0
    new_sheet = False
    test_mode = True
    p_start = 110
    p_end = 120

    return(today_format, headers, offset, attempt, retries, wait, found, yn, do_i_save, total, all_products, prod_count, new_sheet, test_mode, p_start, p_end)


def specs_init():
    global temp_title, prefix, battery_temp, battery_type, mpixels_text
    global mpixels_temp, mpixels, battery_cap, delim, rest, removable
    global month, year, ctext1, ctext2, camera_text, cptext1, cptext2
    global cpu_text, ctext1, ctext2, camera_text, talk_text, wait_text
    global storage_text, ram_text, all_specs, spec_title, spec_specs
    temp_title = prefix = battery_temp = battery_type = mpixels_text = ""
    mpixels_temp = mpixels = battery_cap = delim = rest = removable = ""
    month = year = ctext1 = ctext2 = camera_text = cptext1 = cptext2 = ""
    cpu_text = ctext1 = ctext2 = camera_text = talk_text = wait_text = ""
    storage_text = ram_text = ""
    all_specs = {}
    spec_specs = []
    spec_title = ("01. Διαστάσεις",
                  "02. Βάρος",
                  "03. Χρόνος Ομιλίας",
                  "04. Xρόνος Αναμονής",
                  "05. Οθόνη",
                  "06. Κάρτα μνήμης",
                  "07. Mobile Internet",
                  "08. Ασύρματη επικοινωνία",
                  "09. Camera",
                  "10. Ειδοποιήσεις",
                  "11. MPixels",
                  "12. Εσωτερική μνήμη",
                  "13. Video",
                  "14. Μνήμη RAM",
                  "15. Μπαταρία",
                  "16. Τύπος",
                  "17. Μήνας",
                  "18. Χρόνος",
                  "19. Extra",
                  "20. MMC",
                  "21. JAVA",
                  "22. NFC",
                  "23. GPS",
                  "24. Radio",
                  "25. Fingerprint",
                  "26. Αποσπώμενη μπαταρία",
                  "27. CPU",
                  "28. Ενσωματωμένοι Αισθητήρες",
                  "29. Λειτουργικό Σύστημα"
                  )


try:
    today_format, headers, offset, attempt, retries, wait, found, yn, do_i_save, total, all_products, prod_count, new_sheet, test_mode, p_start, p_end = initialize()
    os.system("title " + "Creating files")
    files_setup(today_format)
    os.system("title " + "Checking txt file")
    no_check()
    page_url = "https://www.e-shop.gr/tilepikoinonies-kinita-smartphones-list?table=TEL&category=%CA%C9%CD%C7%D4%CF+%D4%C7%CB%C5%D6%D9%CD%CF"
    os.system("title " + "Loading soup")
    page_soup = load_soup(page_url, wait, retries)
    os.system("title " + "Getting all items")
    get_all_products(page_url, page_soup, attempt, retries)
    # p_id = 0
    for p_id, product in enumerate(all_products):
        if test_mode:
            if p_start > p_id:
                continue
            elif p_end < p_id:
                break

        dif_found = False
        no_check = False
        # p_id += 1
        title_text = ("Item: " + str(p_id) + "/" + str(len(all_products)
                                                       ) + ". found/yn: " + str(found) + "/" + str(yn))
        os.system("title " + title_text)
        e = 1
        for item in no_check_list:
            if item == product:
                no_check = True
                break
            else:
                no_check = False

        if no_check == True:
            continue

        """ Initializing variables, tables and dics """
        print("Μηδενίζω μεταβλητές...")
        print("")
        specs_init()
        """ Initialization end """

        gr_page_url = 'https://www.e-shop.gr/product?id=' + product  # gr page
        cy_page_url = 'https://www.e-shop.cy/product?id=' + product  # cy page
        print("Προϊόν: " + product + ", " + str(p_id) + "/" + str(len(all_products)))
        print(gr_page_url)
        gr_soup = load_soup(gr_page_url, wait, retries)
        cy_soup = load_soup(cy_page_url, wait, retries)

        gr_specs1, gr_specs2, cy_specs1, cy_specs2 = get_specs(
            gr_soup, cy_soup)

        if len(cy_specs1) < 8:
            for idx in range(0, len(cy_specs1)):
                print("cy_specs1[idx].text.strip: " +
                      cy_specs1[idx].text.strip())
                print("gr_specs1[idx].text.strip: " +
                      gr_specs1[idx].text.strip())
                # if cy_specs1[idx].text.strip() != gr_specs1[idx].text.strip() or cy_specs1[idx].text.strip() == "Ναί" :
                if cy_specs1[idx].text.strip() == "Ναί":
                    print("ΩΩΩΩΩΩΠ βρήκα ναί στο Κυπριακό.")
                    yn += 1
                    ws_write.cell(row=yn + 1, column=1, value=product)
                    ws_write.cell(row=yn + 1, column=2, value="ΝΑΙ/ΟΧΙ")
                    break
        elif len(cy_specs2) == len(gr_specs2):
            compare(gr_specs1, gr_specs2, cy_specs1, cy_specs2)
        elif len(cy_specs2) != len(gr_specs2):
            """ write it on the excel file with a new sheet name """
            print("Βρήκα διαφορές στο πλήθος specs για " + product)
            found += 1
            print("Βρίσκω περιγραφή...")
            desc_text = get_description(gr_soup)
            print("Βρίσκω specs...")
            spec_specs = add_specs(gr_specs1)
            print("Τα 2 σε 1...")
            all_specs = merge_specs(spec_title, spec_specs)
            print("Διορθώνω λάθη...")
            all_specs = spec_fixes(all_specs)
            print("Ταξινομώ...")
            sorted_specs = sort_specs(all_specs)
            title = gr_soup.h1.text.strip()
            print("Γραμμένα τα 'χω...")
            new_sheet = True
            ws_write_product = wb_write.create_sheet(product)
            ws_write_product.cell(row=1, column=1, value="ΤΙΤΛΟΣ")
            ws_write_product.cell(row=1, column=2, value="ΠΕΡΙΓΡΑΦΗ")
            for x, y in sorted_specs.items():
                e += 1
                # print(x + ": " + y)
                # print("writing title")
                ws_write_product.cell(row=e, column=1, value=x)
                # print("writing specs")
                ws_write_product.cell(row=e, column=2, value=y)
                if len(y) >= 50:
                    ws_write_product.cell(row=e, column=3, value=len(y))
            # print("writing description")
            ws_write_product.cell(row=e, column=0, value="ΠΕΡΙΓΡΑΦΗ")
            ws_write_product.cell(row=e, column=1, value=desc_text)
            write_it_down(e, 0)
        print("found = " + str(found))
        print("yn = " + str(yn))
        print("GG")
        print("")
except KeyboardInterrupt:
    try:
        print(" " * 100, "\r")
        input("Διαλλειματάκι;")
        print("")
    except:
        sys.exit(0)
except Exception as exc:
    print("Κάτι δεν πάει καλά.")
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print("Exception: " + str(exc))
    print("Exception type: ", exception_type)
    print("File name: ", filename)
    print("Line number: ", line_number)
    print("")
    try:
        input()
    except KeyboardInterrupt:
        sys.exit(0)

try:
    total = found + yn
    if total == 0:
        print("Δεν βρέθηκαν προϊόντα. Δεν αποθηκεύω.")
        sys.exit(0)
    else:
        if do_i_save == False:
            if total == 1:
                print_text = "Βρήκα 1 προϊόν. Αποθηκεύω?"
            else:
                print_text = "Βρήκα " + \
                    str(total) + " προϊόντα. Αποθηκεύω? "
            answer = input(print_text)
            if answer == "":
                answer = "y"
            if answer != "n" or answer != "N" or answer != "ν" or answer != "Ν":
                write_it_down(new_sheet, 1)
            else:
                print("Δεν αποθηκεύτηκε το αρχείο.")
                sys.exit(0)
        else:
            write_it_down(new_sheet, 1)
except KeyboardInterrupt:
    sys.exit(0)
except Exception as exc:
    print("Κάτι δεν πάει καλά.")
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print("Exception: " + str(exc))
    print("Exception type: ", exception_type)
    print("File name: ", filename)
    print("Line number: ", line_number)
    print("")
    try:
        input()
    except KeyboardInterrupt:
        sys.exit(0)

sys.exit(0)