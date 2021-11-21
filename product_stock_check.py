import sys


def ti_paizei(show_version):
    # product_stock_check.py
    # # #
    # Έλεγχος στο αρχείο αλλαγής τιμών, φύλλο κατεβασμένα για έλεγχο του στοκ.
    # # #
    # Current Version 1 beta
    #######
    version = "Version 1 beta"
    print(version)


try:
    from bs4 import BeautifulSoup as soup
    from time import sleep as nani
    from time import time
    from random import randint as dose
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    import requests
    import os
    import string
    import subprocess
    import psutil
except KeyboardInterrupt:
    sys.exit(1)
except Exception as exc:
    import sys
    print("Κάτι πάθαμε κατά το import.")
    print(str(exc))
    sys.exit(0)


def initialize():
    print("Αρχικοποίηση παραμέτρων...")
    test_run = 0
    attempt = 0  # how many attempts to re-read the url in case of failure
    e = 2  # will add up in case of exceptions
    retries = 10
    wait = 3
    headers = {
        'User-Agent': "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.27 Safari/537.17"}
    office_run = False
    convert_xl = True
    show_version = False
    run_gr = False
    run_cystock = True
    success = False
    est_total = 0
    est_rem = 0
    formatted_est = ""
    print("Done")
    return test_run, attempt, e, retries, wait, headers, office_run, convert_xl, show_version, run_gr, run_cystock, success, est_total, est_rem, formatted_est


def get_start_time():
    global start_time, start_date
    start = datetime.now()
    start_date = start.strftime("%d-%m-%Y")
    start_time = start.strftime("%H:%M:%S")
    print("Εκκίνηση:", start_date, ",", start_time)
    start_time = time()
    print("")


def convert_ods_xls(ods_file, xl_file, write_path, full_ods):
    EXE = 'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
    print("Μετατροπή του: " + ods_file + " σε " +
          xl_file + ". Θα πάρει λίγη ωρίτσα...")
    subprocess.run([EXE, '--convert-to', 'xlsx',
                   full_ods, '-outdir', write_path])
    print("Έτοιμος.")


def soffice_runs():
    print("Για να δούμε τρέχει το Libre?")
    processName = "soffice.exe"
    attempt = 0
    while attempt < 3:
        for proc in psutil.process_iter():
            #  print(str(proc.name().lower()))
            if processName.lower() in proc.name().lower():
                print("Ώπα τρέχει. Περιμένω 1 λεπτό να το κλείσεις και ξαναπροσπαθώ.")
                nani(10)
                attempt += 1
            else:
                pass
        break

    if attempt == 3:
        print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
        print("Κλέισε το office και προσπάθησε πάλι.")
        input()
        sys.exit(0)
    else:
        print("Είσαι 'ντάξ'.")


def get_file_info(full_file):
    file_stats = os.stat(full_file)
    file_modified = file_stats.st_mtime

    return(file_modified)


def set_read_files():
    global write_path, ods_file, wb
    try:
        if os.path.exists("Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product"):
            write_path = ("Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product")
        elif os.path.exists("Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product"):
            write_path = ("Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product")
        os.chdir(write_path)
        if os.path.exists('Αλλαγή τιμών.ods'):
            ods_file = ('Αλλαγή τιμών.ods')  # path to ods read file
        else:
            ods_file = ('Αλλαγή τιμών_2.ods')  # path to ods read file

        xl_file = 'Αλλαγή τιμών.xlsx'
        full_ods = os.path.join(write_path, ods_file)
        full_xl = os.path.join(write_path, xl_file)
        ods_modified = datetime.fromtimestamp(get_file_info(full_ods))
        xls_modified = datetime.fromtimestamp(get_file_info(full_xl))

        if office_run:
            soffice_runs()

        if xls_modified < ods_modified and convert_xl:
            convert_ods_xls(ods_file, xl_file, write_path, full_ods)

        print("")
        print(
            "Προσπαθώ να ανοίξω το: " + full_xl + ". Θα πάρει λίγη ωρίτσα...")
        wb = load_workbook(full_xl)
        print("Τα καταφέραμε.")
        print("")

    except Exception as exc:
        print(str(exc))
        print("Δεν βρίσκω το αρχείο " + full_ods + " ή δεν ανοίγει.")
        print("")
        sys.exit()


def set_sheets():
    global ac_row, ac_col, col_index, row_index, sheet, sheets, sheet_index, sheet_name

    # Επιλογή φύλλου
    sheet_list = []
    sheets = wb.sheetnames
    print("Μαζεύω τα φύλλα... υπομονή.")
    for i in range(0, len(sheets)):
        sheet_list.append(sheets[i])

    sheet_index = 1
    sheet = wb[sheet_list[sheet_index - 1]]
    sheet_name = sheet_list[sheet_index - 1]
    print("")

    # Μέτρημα και επιλογή γραμμών

    colcount = sheet.max_column
    ac_col = 0

    for i in range(1, colcount):
        col_value = str(sheet.cell(row=1, column=i).value)
        if col_value == "" or col_value is None or col_value == "None":
            break
        else:
            print('Στήλη ' + str(i) + ': ' + col_value)
            ac_col += 1

    col_index = 1
    print("")
    print("Επιλέγω στήλη: " + str(col_index) + " - " +
          str(sheet.cell(row=1, column=col_index).value))
    print("")

    # Μέτρημα και επιλογή γραμμών

    rowcount = sheet.max_row
    ac_row = 0
    for i in range(1, rowcount):
        row_value = str(sheet.cell(row=i, column=1).value)
        if row_value == "" or row_value is None or row_value == "None":
            break
        else:
            ac_row += 1

    row_index = 1
    print("")
    print("Ξεκινάμε από γραμμή: " + str(row_index))
    print("")


def set_write_files():
    global write_file, alt_write_file, wb_write, ws_write
    os.chdir(write_path)
    write_file = ("product_stock_check.xlsx")  # name of xls write file
    alt_write_file = ("product_stock_check_alt.xlsx")
    wb_write = Workbook()
    ws_write = wb_write.active
    ws_write.title = "stockcheck"
    ws_write.cell(row=1, column=1, value="ΚΩΔΙΚΟΣ")
    ws_write.cell(row=1, column=2, value="ΤΙΤΛΟΣ")
    ws_write.cell(row=1, column=3, value="ΤΙΜΗ CY")
    ws_write.cell(row=1, column=4, value="ΤΙΜΗ GR")
    ws_write.cell(row=1, column=5, value="STOCK CY")

    # create a list with all english letters representing column names
    alphabet = list(string.ascii_uppercase)
    for w in range(1, 6):
        c_name = alphabet[w - 1]  # column name
        c_value = ws_write.cell(row=1, column=w).value  # column header text
        c_length = len(c_value)  # column length
        # if c_value.find("ΚΩΔΙΚΟΣ") >= 0 or c_value.find("LINK") >= 0:
        if c_value.find("ΚΩΔΙΚΟΣ") >= 0:
            c_length = 12
        elif c_value.find("ΤΙΤΛΟΣ") >= 0:
            c_length = 24
        else:
            c_length = 8
        ws_write.column_dimensions[c_name].width = c_length


def load_soup(page, wait, retries):
    # print("Μέσα στη σούπα.")
    attempt = 0
    while attempt < retries:
        try:
            result = requests.get(page, headers=headers)
            webpage = result.content
            page_soup = soup(webpage, "html5lib")
            # print(headers)
            break
            # print("Έξω από τη σούπα.")
            # print("")
        except Exception as exc:
            print("")
            print("Στο φόρτωμα της σελίδας, πέσαμε πάνω στο:")
            print(str(exc))
            print("Ξαναπροσπαθώ σε " + str(retries) + ".")
            nani(wait)
            attempt += 1
    if attempt == retries:
        print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
        input()
        sys.exit(0)

    return(page_soup)


def get_cy_price(page_soup):
    # global cy_price, cy_title
    cy_title = page_soup.h1.text
    cy_price_soup = page_soup.findAll("span", {"class": "web-price-value-new"})
    if len(cy_price_soup) == 0:
        cy_price_text = "ΕΞΑΝΤΛΗΜΕΝΟ"
    else:
        cy_price_text = cy_price_soup[0].text.replace("\xa0€", "")

    try:
        cy_price = float(cy_price_text)
    except Exception as exc:
        cy_price = cy_price_text

    return(cy_price, cy_title)


def get_gr_price(page_soup):
    gr_title = page_soup.h1.text
    gr_price_soup = page_soup.findAll("span", {"class": "web-price-value-new"})
    if len(gr_price_soup) == 0:
        gr_price_text = "ΕΞΑΝΤΛΗΜΕΝΟ"
    else:
        gr_price_text = gr_price_soup[0].text.replace("\xa0€", "")

    try:
        gr_price = float(gr_price_text)
    except Exception as exc:
        gr_price = gr_price_text

    return(gr_price, gr_title)


def get_cy_details(page_soup):
    avail = page_soup.find("td", {
                           "style": "text-align:left;padding:5px 0 2px 5px;color:#4f4f4f;font-family:Tahoma;font-size:14px;font-weight:bold;"})
    avail_text = avail.text
    if avail_text.find('ËÅÌ:') >= 0:
        avail_text_lim = avail_text[avail_text.find(
            'ËÅÌ: ')+5:avail_text.find('ËÅÕ: ')-1].strip()
        avail_text_nic = avail_text[avail_text.find(
            'ËÅÕ: ')+5:avail_text.find('ËÁÑ: ')-1].strip()
        avail_text_lar = avail_text[avail_text.rfind(': ')+1:].strip()
    else:
        avail_text_lim = avail_text[avail_text.find(
            'ΛΕΜ: ')+5:avail_text.find('ΛΕΥ: ')-1].strip()
        avail_text_nic = avail_text[avail_text.find(
            'ΛΕΥ: ')+5:avail_text.find('ΛΑΡ: ')-1].strip()
        avail_text_lar = avail_text[avail_text.rfind(': ')+1:].strip()
        avail_total = int(avail_text_nic) + \
            int(avail_text_lim) + int(avail_text_lar)

    return avail_total


def write_results(e):
    font_green = Font(color="8FCE00")
    font_red = Font(color="F44336")

    ws_write.cell(row=e, column=1, value=eshop_code)
    ws_write.cell(row=e, column=2, value=cy_title)
    if cy_price != "ΕΞΑΝΤΛΗΜΕΝΟ":
        ws_write.cell(row=e, column=3).font = font_red
    else:
        ws_write.cell(row=e, column=3).font = font_green
    ws_write.cell(row=e, column=3, value=cy_price)
    ws_write.cell(row=e, column=4, value=gr_price)
    ws_write.cell(row=e, column=5, value=total_stock)

    e += 1
    return e


def write_it_down(e, write_file):
    changed = e - 2
    if e > 2:
        success = False
        try:
            wb_write.save(write_file)
            # print(str(e))
            # print("Βρέθηκαν αλλαγές: " + str(changed))
            print(write_file + " σώθηκε στο " + write_path)
            success = True
        except Exception as exc:
            exception_type, exception_object, exception_traceback = sys.exc_info()
            filename = exception_traceback.tb_frame.f_code.co_filename
            line_number = exception_traceback.tb_lineno
            print("Εξαίρεση:        " + str(exc))
            print("Τύπος εξαίρεσης: ", exception_type)
            input("Μήπως είναι ανοιχτό το αρχείο; Κλείσε και προσπάθησε πάλι...")
    else:
        print("Δεν βρέθηκαν αλλαγές δεν σώζω αρχείο.")
        success = True
    return(success)


def write_it_down_old(e, write_file):
    changed = e - 2
    the_now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if e > 2:
        success = False
        try:
            wb_write.save(write_file)
            success = True
        except Exception as exc:
            print("Exception: " + str(exc))
            write_file = alt_write_file
            wb_write.save(write_file)
            success = True
        finally:
            if success is False:
                print(
                    "Και τα 2 αρχεία είναι μάλλον ανοιχτά. Προσπαθώ να σώσω σε νέο αρχείο.")
                write_file = "times_dif_" + dose(1, 100) + "_" + write_file
                wb_write.save(write_file)
            else:
                print("-----------------")
                print("Βρέθηκαν αλλαγές: " + str(changed))
                print(write_file + " σώθηκε στο " + write_path)
                print("-----------------")
    else:
        print("Δεν βρέθηκαν αλλαγές δεν σώζω αρχείο.")
    return(success)


def time_it(start_time, c_run, t_runs):
    time_diff = time() - start_time
    aver = round(time_diff / c_run, 2)
    est_total = round(aver * t_runs, 2)
    est_rem = round(est_total - time_diff, 2)

    # calculate hours, minutes, seconds and format them fo a nice display
    rem_mins = int(est_rem / 60)
    rem_hours = int(rem_mins / 60)
    rem_seconds = int(est_rem - (rem_mins * 60))
    formatted_est = str(rem_hours).zfill(2) + ":" + str(rem_mins).zfill(2) + ":" + str(rem_seconds).zfill(2)
    return(est_total, est_rem, formatted_est)


try:
    test_run, attempt, e, retries, wait, headers, office_run, convert_xl, show_version, run_gr, run_cystock, success, est_total, est_rem, formatted_est = initialize()
    if show_version:
        ti_paizei()
    get_start_time()
    os.system("title " + "Ξεκινήσαμε. Άνοιγμα αρχείων...")
    set_read_files()
    os.system("title " + "Επιλογές...")
    set_sheets()
    set_write_files()
    print(str(row_index) + "/" + str(col_index))
    for r in range(row_index + 1, ac_row + 1):
        cell_value = str(sheet.cell(row=r, column=col_index).value)
        if cell_value.strip() == "" or cell_value is None or cell_value == "None":
            break
        else:
            current_row = r - 1
            last_row = ac_row - 1
        if current_row < last_row - 1:
            print_text = "@" + str(current_row) + ". Απομένουν: " + \
                str(ac_row - r) + "/" + str(last_row) + " γραμμές."
        elif current_row == last_row - 1:
            print_text = "@" + str(current_row) + ". Απομένει 1 γραμμή."
        else:
            print_text = "Τελευταία γραμμή."
        if est_total != 0 and est_rem != 0:
            print_text += " ETA: " + str(formatted_est)
        print(print_text)
        eshop_code = cell_value.strip()
        os.system("title " + print_text + " - " + eshop_code)
        cy_url = "https://www.e-shop.cy/product?id=" + eshop_code
        gr_url = "https://www.e-shop.gr/s/" + eshop_code
        cy_soup = load_soup(cy_url, wait, retries)
        cy_price, cy_title = get_cy_price(cy_soup)
        print("Κωδικός:  " + eshop_code)
        print("Τίτλος:   " + cy_title)
        print("CY Price: " + str(cy_price))
        if run_gr:
            gr_soup = load_soup(gr_url, wait, retries)
            gr_price, gr_title = get_gr_price(gr_soup)
        else:
            gr_price = gr_title = "-"

        if run_cystock:
            try:
                total_stock = get_cy_details(cy_soup)
            except Exception as exc:
                total_stock = "-"
        else:
            total_stock = "-"

        e = write_results(e)
        est_total, est_rem, formatted_est = time_it(start_time, current_row, last_row)

except KeyboardInterrupt:
    print("")
    print("OK κατάλαβα. Διαλλειματάκι... ")
    print("")
except Exception as exc:
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print("Exception: " + str(exc))
    print("Exception type: ", exception_type)
    print("File name: ", filename)
    print("Line number: ", line_number)


try:
    while success is False:
        success = write_it_down(e, write_file)

except Exception as exc:
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print("Exception: " + str(exc))
    print("Exception type: ", exception_type)
    print("File name: ", filename)
    print("Line number: ", line_number)

input()
sys.exit(0)
