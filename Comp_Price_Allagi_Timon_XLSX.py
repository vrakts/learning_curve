def ti_paizei(show_version):
 # Comp_Price_Allagi_Timon_XLSX.py
 # Έλεγχος αν η τιμή στο αρχείο Αλλαγή τιμών έχει αλλάξει σε CY, GR και ανταγωνιστές
 # Ο στόχος είναι να ελέγχει αν κάποιος ανταγωνιστής έχει χαμηλότερη τιμή και να το καταγράφει.
 ######
 ### To do
 # - Να τρέχει το ods to xl αν το modified date είναι διαφορετικό. --- DONE
 # - Hyperlinks: --- DONE
 #    wbook.active['A8'].hyperlink = "http://www.espn.com"
 #    wbook.active['A8'].value = 'ESPN'
 #    wbook.active['A8'].style = "Hyperlink"
 # - Προσθήκη χρώματος. --- DONE
 # - Να γράφει ή να κοκκινίζει την τιμή CY και GR μόνο αν έχουν αλλάξει. --- DONE κοκκινίζει το GR
 # - Διόρθωση των counter.
 # - Έλεγχος ότι του URL του ανταγωνιστή είναι σωστό.
 # = Έλεγχος διαθεσιμότητας και καταγραφή του κάπως (πορτοκαλί χρώμα ίσως).
 # - Συμμάζεμα definitions.
 if show_version == True:
  print("Version: 1.0 Beta")
 else:
  pass

import sys

try :
 from bs4 import BeautifulSoup as soup  # import the BeatifulSoup function from bs4 as soup
 from time import sleep as nani
 from random import randint as dose
 from datetime import datetime
 from selenium import webdriver
 from selenium.webdriver import ChromeOptions
 from openpyxl import Workbook
 from openpyxl import load_workbook
 from openpyxl.styles import Font
 # from selenium.webdriver.common.keys import Keys
 # from selenium.webdriver.chrome.options import Options  
 # import re
 # import ezodf
 import requests
 import os
 # import xlwt
 # import openpyxl
 import string
 import subprocess
 import psutil
except KeyboardInterrupt :
 sys.exit(1)
except Exception as exc :
 import sys
 print("Κάτι πάθαμε κατά το import.")
 print(str(exc))
 sys.exit(0)

def get_start_time() :
 global start_time, start_date
 start = datetime.now()
 start_date = start.strftime("%d-%m-%Y")
 start_time = start.strftime("%H:%M:%S")
 print("Εκκίνηση: " + start_date)
 print("")

def convert_ods_xls(ods_file, xl_file, write_path, full_ods):
 EXE = 'C:\Program Files\LibreOffice\program\soffice.exe'
 print("Μετατροπή του: " + ods_file + " σε " + xl_file + ". Θα πάρει λίγη ωρίτσα...")
 subprocess.run([EXE, '--convert-to', 'xlsx', full_ods, '-outdir', write_path])
 print("Έτοιμος.")

def get_file_info(full_file):
 file_stats = os.stat(full_file)
 file_modified = file_stats.st_mtime
 
 return(file_modified)

def set_read_files() :
 global write_path, ods_file, wb
 try :
  if os.path.exists("Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product") == True :
   write_path = ("Z:\\OneDrive\\eShop Stuff\\PRODUCT\\Product")
  elif os.path.exists("Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product") == True :
   write_path = ("Y:\\OneDrive\\eShop Stuff\\PRODUCT\\Product")
  os.chdir(write_path)
  if os.path.exists('Αλλαγή τιμών.ods') :
   ods_file = ('Αλλαγή τιμών.ods')  # path to ods read file
  else :
   ods_file = ('Αλλαγή τιμών_2.ods')  # path to ods read file
  
  xl_file = 'Αλλαγή τιμών.xlsx'
  full_ods = os.path.join(write_path, ods_file)
  full_xl = os.path.join(write_path, xl_file)
  ods_modified = datetime.fromtimestamp(get_file_info(full_ods))
  xls_modified = datetime.fromtimestamp(get_file_info(full_xl))

  if office_run == True:
   soffice_runs()
  
  if convert_xl == True and xls_modified < ods_modified:
   convert_ods_xls(ods_file, xl_file, write_path, full_ods)
  
  print("")
  print("Προσπαθώ να ανοίξω το: " + full_xl + ". Θα πάρει λίγη ωρίτσα...")
  wb = load_workbook(full_xl)
  print("Τα καταφέραμε.")
  print("")
  
 except Exception as exc :
  print(str(exc))
  print("Δεν βρίσκω το αρχείο " + full_ods + " ή δεν ανοίγει.")
  print("")
  sys.exit()

def set_sheets() :
 global ac_row, ac_col, col_index, row_index, sheet, sheets, sheet_index, sheet_name
 
 # Επιλογή φύλλου
 sheet_list = []
 sheets = wb.sheetnames
 print("Μαζεύω τα φύλλα... υπομονή.")
 for i in range(0, len(sheets)) :
  sheet_list.append(sheets[i])
 
 for i in range(0, len(sheet_list)) :
  print('Φύλλο ' + str(i + 1) + ': ' + sheet_list[i])
 
 answer = 'Διάλεξε φύλλο: '
 sheet_index = input(answer)
 # print("Τύπος: " + str(type(sheet_index)))
 # print("sheet_index: " + sheet_index)
 # print("is empty? " + str(sheet_index == ""))
 if sheet_index == "" :
  sheet_index = 1
 else :
  sheet_index = int(sheet_index)
 
 # print("Τύπος: " + str(type(sheet_index)))
 # print("sheet_index: " + str(sheet_index))
 # print("is empty? " + str(sheet_index == ""))
 
 print("Επιλέγω: " + sheet_list[sheet_index - 1])
 sheet = wb[sheet_list[sheet_index - 1]]
 sheet_name = sheet_list[sheet_index - 1]
 print("")
 
 # Μέτρημα και επιλογή στηλών
 
 colcount = sheet.max_column
 ac_col = 0
 
 for i in range(1, colcount) :
  col_value = str(sheet.cell(row=1, column=i).value)
  if col_value == "" or col_value == None or col_value == "None":
   break
  else:
   print('Στήλη ' + str(i) + ': ' + col_value)
   ac_col += 1
 
 answer = 'Διάλεξε στήλη: '
 col_index = input(answer)
 if col_index == "" :
  col_index = 1
 else :
  col_index = int(col_index)
 print("")
 print("Επιλέγω στήλη: " + str(col_index) + " - " + str(sheet.cell(row=1, column=col_index).value))
 print("")
 
 # Μέτρημα και επιλογή γραμμών
 
 rowcount = sheet.max_row
 ac_row = 0
 for i in range(1, rowcount):
  row_value = str(sheet.cell(row=i, column=1).value)
  if row_value == "" or row_value == None or row_value == "None":
   break
  else:
   ac_row += 1
 
 answer = 'Αρχική γραμμή: '
 row_index = input(answer)
 if row_index == "" :
  row_index = 1
 else :
  row_index = int(row_index)
 print("")
 print("Ξεκινάμε από γραμμή: " + str(row_index))
 print("")

def set_write_files(sheet_name) :
 global write_file, alt_write_file, wb_write, ws_write
 os.chdir(write_path)
 write_file = ("times_dif_" + str(sheet_index) + "-" + str(sheet_name) + "_" + start_date + ".xlsx")  # name of xls write file
 alt_write_file = ("times_dif_alt_" + str(sheet_index) + "-" + str(sheet_name) + "_" + start_date + ".xlsx")  # alternate name of xls write file
 wb_write = Workbook()
 ws_write = wb_write.active
 ws_write.title = "Diffs"
 # ws_write.cell(row=1, column=10).hyperlink="http://gvrakas.com"
 ws_write.cell(row=1, column=1, value="ΚΩΔΙΚΟΣ")
 ws_write.cell(row=1, column=2, value="ΤΙΤΛΟΣ")
 ws_write.cell(row=1, column=3, value="ΤΙΜΗ CY")
 ws_write.cell(row=1, column=4, value="ΤΙΜΗ GR")
 ws_write.cell(row=1, column=5, value="ST PRICE")
 ws_write.cell(row=1, column=6, value="ST LINK")
 ws_write.cell(row=1, column=7, value="PUB PRICE")
 ws_write.cell(row=1, column=8, value="PUB LINK")
 ws_write.cell(row=1, column=9, value="EL PRICE")
 ws_write.cell(row=1, column=10, value="EL LINK")
 ws_write.cell(row=1, column=11, value="KOT PRICE")
 ws_write.cell(row=1, column=12, value="KOT LINK")
 ws_write.cell(row=1, column=13, value="BIO PRICE")
 ws_write.cell(row=1, column=14, value="BIO LINK")
 ws_write.cell(row=1, column=15, value="SIN PRICE")
 ws_write.cell(row=1, column=16, value="SIN LINK")
 ws_write.cell(row=1, column=17, value="CUS PRICE")
 ws_write.cell(row=1, column=18, value="CUS LINK")
 
 ws_write.column_dimensions['A'].width = 12
 ws_write.column_dimensions['B'].width = 24
 ws_write.column_dimensions['C'].width = 8
 ws_write.column_dimensions['D'].width = 8

 
 alphabet = list(string.ascii_uppercase)  # create a list with all english letters representing column names
 for w in range(1, 19):
  c_name = alphabet[w - 1]  # column name
  c_value = ws_write.cell(row=1, column=w).value  # column header text
  c_length = len(c_value)  # column length
  # if c_value.find("ΚΩΔΙΚΟΣ") >= 0 or c_value.find("LINK") >= 0:
  if c_value.find("ΚΩΔΙΚΟΣ") >= 0:
   c_length = 12
  elif c_value.find("ΤΙΤΛΟΣ") >= 0:
   c_length = 24
  else:
   c_length = 8
  
  ws_write.column_dimensions[c_name].width = c_length

def load_soup(page, wait, retries) :
 # print("Μέσα στη σούπα.")
 attempt = 0
 while attempt < retries :
  try :
   result = requests.get(page, headers = headers)
   webpage = result.content
   page_soup = soup(webpage, "html5lib")
   # print(headers)
   break   
   # print("Έξω από τη σούπα.")
   # print("")
  except Exception as exc :
   print("")
   print("Στο φόρτωμα της σελίδας, πέσαμε πάνω στο:")
   print(str(exc))
   print("Ξαναπροσπαθώ σε " + str(retries)+ ".")
   nani(wait)
   attempt += 1
 if attempt == retries :
  print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
  input()
  sys.exit(0)
 
 return(page_soup)

def margin_check(page_soup) :
 margin_exist = page_soup.findAll("font", {"style" : "color:#ff9933;font-weight:bold;font-size:9px;font-family:arial black;"})
 stock_exist = page_soup.findAll("font", {"style" : "color:#ff0000;font-weight:bold;font-size:9px;font-family:arial black;"})
 if len(stock_exist) != 0 :  # if stock_exist not empty then stock sign exists
  margin = "STOCK"
 elif len(margin_exist) == 0 :  # if margin_exist is empty then margin sign doesn't exist and should be corrected
  margin = "MARGIN"
 else :  # if both stock sign doesn't exist and high margin exist then no changes
  margin = "-"

def get_cy_price(page_soup) :
 # global cy_price, cy_title
 cy_title = page_soup.h1.text
 cy_price_soup = page_soup.findAll("span", {"class" : "web-price-value-new"})
 if len(cy_price_soup) == 0 :
  cy_price_text = "ΕΞΑΝΤΛΗΜΕΝΟ"
 else : 
  cy_price_text = cy_price_soup[0].text.replace("\xa0€", "")
 
 try :
  cy_price = float(cy_price_text)
 except :
  cy_price = cy_price_text
 
 return(cy_price, cy_title)

def get_gr_price(page_soup) :
 gr_title = page_soup.h1.text
 gr_price_soup = page_soup.findAll("span", {"class" : "web-price-value-new"})
 if len(gr_price_soup) == 0 :
  gr_price_text = "ΕΞΑΝΤΛΗΜΕΝΟ"
 else : 
  gr_price_text = gr_price_soup[0].text.replace("\xa0€", "")
 
 try :
  gr_price = float(gr_price_text)
 except :
  gr_price = gr_price_text
 
 return(gr_price, gr_title)

def soffice_runs():
 print("Για να δούμε τρέχει το Libre?")
 processName = "soffice.exe"
 attempt = 0
 while attempt < 3 :
  for proc in psutil.process_iter():
  #  print(str(proc.name().lower()))
   if processName.lower() in proc.name().lower():
    print("Ώπα τρέχει. Περιμένω 1 λεπτό να το κλείσεις και ξαναπροσπαθώ.")
    nani(10)
    attempt += 1
   else:
    pass
  break
  
 if attempt == 3 :
  print("Προσπάθησα " + str(attempt) + " φορές και δεν τα κατάφερα.")
  print("Κλέισε το office και προσπάθησε πάλι.")
  input()
  sys.exit(0)
 else:
  print("Είσαι 'ντάξ'.")

def st_stuff(link) :  # supposedly fixed.
 product = load_soup(link, wait, retries)
 
 if product.find('div', {'class': 'listing-details-column large-stephanis-card-price large-single'}) :
  display_price = product.find('div', {'class': 'listing-details-column large-stephanis-card-price large-single'}).div.text.replace('€', '').strip()
 elif product.find('div', {'class': 'listing-details-heading large-now-price with-sale'}) :
  display_price = product.find('div', {'class': 'listing-details-heading large-now-price with-sale'}).text.replace('€', '').strip()
 else :
  display_price = "-"
 
 try:
  price = float(display_price.replace(',', ''))
 except:
  price = "ΕΞΑΝΤΛΗΜΕΝΟ"
 
 # print('price:    ' + str(price))
 return(price)

def pub_stuff(link) :  # supposedly fixed using selenium.
 # product = load_soup(link, wait, retries)
 driver.get(link)
 nani(5)
 product = soup(driver.page_source, features = "lxml")
 # driver.close() 
 
 if product.find('div', {'class': 'teaser--product-final-price large sell-price'}):
  full_price = product.find('div', {'class': 'teaser--product-final-price large sell-price'}).text.strip()
  decimals_price = product.find('div', {'class': 'teaser--product-final-price large sell-price'}).span.text.strip()
  int_price = full_price.replace(decimals_price, "").replace(".","")
  display_price = int_price + "." + decimals_price.replace("€", "")
  price = float(display_price)
 else:
  price = "ΕΞΑΝΤΛΗΜΕΝΟ"
 
 # driver.close()
 # print('price:    ' + str(price))
 return(price)

def el_stuff(link) :  # supposedly fixed.
 product = load_soup(link, wait, retries)
  
 if product.find('div', {'class': 'single-product-prices'}).ins :
  display_price = product.find('div', {'class': 'single-product-prices'}).ins.text.replace('€', '').strip()
 else :
  display_price = product.find('div', {'class': 'single-product-prices'}).h2.text.replace('€', '').strip()
 
 # elif product.find('span', {'class' : 'listing-product-price listing-product-price--without-loyalty'}) :
 # display_price = product.find('span', {'class' : 'listing-product-price listing-product-price--without-loyalty'}).text.replace('€', '').strip()
 # else :
 # display_price = product.find('div', {'class' : 'listing-product-price listing-product-price--rows-layout'}).text.replace('€', '').strip()
 
 price = float(display_price.replace(',', ''))
 # print('price:    ' + str(price))
 return(price)

def bio_stuff(link) :  # supposedly fixed using selenium.
 
 driver.get(link)
 nani(2)
 product = soup(driver.page_source, features = "lxml")
 
 price_soup = product.find('div', {'class': 'product-prices-wrapper'})

 if price_soup.find('div', {'class': 'price loyalty'}) :
  display_price = price_soup.find('div', {'class': 'price loyalty'}).h3.text.replace('€', '').strip()
 elif price_soup.find('div', {'class': 'price periodic'}) :
  display_price = price_soup.find('div', {'class': 'price periodic'}).h3.text.replace('€', '').strip()
 elif price_soup.find('div', {'class': 'retail-price'}) :
  display_price = price_soup.find('div', {'class': 'retail-price'}).h3.text.replace('€', '').strip()
 else :
  display_price = price_soup.find('div', {'class': 'price regular'}).h3.text.replace('€', '').strip()
 price = float(display_price.replace(',', ''))
 
 # driver.close()
 # print('price:    ' + str(price))
 return(price)

def sin_stuff(link) :  # supposedly fixed.
 product = load_soup(link, wait, retries)
 
 display_price = product.find('span', {'class' : 'ty-list-price ty-nowrap'}).text.replace("€", "").replace("inc. VAT", "").strip()
 price = float(display_price.replace(',', ''))
 
 # print('price:    ' + str(price))
 return(price)

def cus_stuff(link) :  # supposedly fixed.
 product = load_soup(link, wait, retries)
 
 if product.find('span', {'class': 'ty-price'}):
  full_price = product.find('span', {'class': 'ty-price'}).text.strip().replace("€", "")
  decimals_price = product.find('span', {'class': 'ty-price'}).sup.text.strip()
  int_price = full_price.replace(decimals_price, "").replace(".","")
  display_price = int_price + "." + decimals_price.replace("€", "")
  price = float(display_price.replace(',', ''))
 else:
  price = "ΕΞΑΝΤΛΗΜΕΝΟ"
 
 # print('price:    ' + str(price))
 return(price)

def kot_stuff(link) :  # supposedly fixed.
 product = load_soup(link, wait, retries)
 
 if product.text.find('κερδίζεις') >= 0 :
  price = product.find('div', {'class' : 'price'}).text.strip().replace('\n', '').replace('\t', '')[4:]
  # init_price = price[:price.find('€')]
  discount_price = price[price.find('€') + 1:price.find('ΤΙΜΗ')]
  price = float(discount_price.replace(',', ''))
 else :
  price = product.find('div', {'class' : 'price'}).text.strip().replace('\n', '').replace('\t', '')
  init_price = price[1:price.find('ΤΙΜΗ')]
  discount_price = "-"
  price = float(init_price.replace(',', ''))
 
 # print('price:    ' + str(price))
 return(price)

def initialize():
 print("Αρχικοποίηση παραμέτρων...")
 test_run = 0
 attempt = 0  # how many attempts to re-read the url in case of failure
 e = 2  # will add up in case of exceptions
 retries = 10
 wait = 3
 headers = {'User-Agent': "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.27 Safari/537.17"}
 cookies = {'language': 'en', '_myPublicID': 'G-ed9dbb98-2434-290f-2b01-af19a2e28e53', '_pic': '4257825144', 'JSESSIONID': 'O6FeVAjR0CFrOlopHaLplYcR.node2', 'roid': 'o014661427', 'snalyticsi': '17b2b49c11348ecf22f93d622eb14658'}
 office_run = False
 convert_xl = True
 show_version = False
 ti_paizei(show_version)
 print("Done")
 return test_run, attempt, e, retries, wait, headers, cookies, office_run, convert_xl

def chrome_init():
 chrome_options = ChromeOptions()
 chrome_options.add_argument("--headless")
 chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
 chrome_options.add_argument("--log-level=OFF")
 try:
  print("Ξεκινάω τον Chrome Driver για τους δύσκολους...")
  driver = webdriver.Chrome("C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe", options = chrome_options)
  chrome_drv = True
  print("Chrome Driver ΟΚ...")
 except Exception as exc:
  print("Φαίνεται πως δεν ξεκίνησε ο Chrome Driver. Κάποιοι ανταγωνιστές δεν θα δουλέψουν σήμερα...")
  print("Error: " + str(exc))
  chrome_drv = False
 
 print("")
 return chrome_options, driver, chrome_drv

def was_is(was_price, is_price):
 print("Was:      " + was_price + ", is: " + str(is_price))

def comp_init():
 comp_name = comp_link = was_price = st_price = st_link = pub_price = pub_link = el_price = el_link = kot_price = kot_link = bio_price = bio_link = si_price = si_link = cus_price = cus_link = "-"
 comp_wins = False
 return comp_name, comp_link, was_price, st_price, st_link, pub_price, pub_link, el_price, el_link, kot_price, kot_link, bio_price, bio_link, si_price, si_link, cus_price, cus_link, comp_wins

def write_results(e, comp_list):
 col_write = 1
 font_green = Font(color="8FCE00")
 font_red = Font(color="F44336")
 for comp_value in comp_list:
  column_title  = ws_write.cell(row=1, column=col_write).value
  ws_write.cell(row=e, column=col_write, value=comp_value)
  if column_title.find("LINK") >= 0:
   ws_write.cell(row=e, column=col_write).hyperlink = comp_value
   ws_write.cell(row=e, column=col_write).style = "Hyperlink"
  
  comp_cell = ws_write.cell(row=e, column=col_write)
  cy_cell = ws_write.cell(row=e, column=3)
  gr_cell = ws_write.cell(row=e, column=4)
  cy_value = cy_cell.value
  gr_value = gr_cell.value
  
  # if column_title.find("ΤΙΜΗ GR") >= 0 and gr_value > cy_value:
  if column_title.find("ΤΙΜΗ GR") >= 0 or col_write == 4:
   if type(cy_value) == float and type(gr_value) == float and cy_value < gr_value:
    print("Βρήκα ψηλότερη στο GR " + str(gr_value) + " vs " + str(cy_value))
    print("")
    gr_cell.font = font_red
    # print(gr_cell.font.color.rgb)
   else:
    gr_cell.font = font_green
    # print(gr_cell.font.color.rgb)
  elif column_title.find("PRICE") >= 0:
   if type(comp_value) == float and type(cy_value) == float and comp_value < cy_value:
    comp_cell.font = font_red
   else:
    comp_cell.font = font_green
  else:
   pass
  
  col_write += 1
 
 e += 1
 
 return e

def write_it_down(e, index, name) :
 changed = e - 2
 the_now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
 filename_end = str(index) + "-" + str(name).replace(" ", "-") + "-" + the_now + ".xlsx"
 write_file = "times_dif_" + filename_end
 alt_write_file = "times_dif_alt_" + filename_end
 print("Όνομα αρχείου: " + write_file)
 print("Εναλλακτικά:   " + alt_write_file)
 if e > 2:
  success = False
  try :
   wb_write.save(write_file)
   success = True
  except :
   wb_write.save(alt_write_file)
   success = True
  finally :
   if success == False :
    print("Και τα 2 αρχεία είναι μάλλον ανοιχτά. Προσπαθώ να σώσω σε νέο αρχείο.")
    write_file = "times_dif_" + dose(1, 100) + "_" + filename_end
    wb_write.save(write_file)
   else :
    print(str(e))
    print("Βρέθηκαν αλλαγές: " + str(changed))
    print(write_file + " σώθηκε στο " + write_path)
   #  sys.exit(write_file + " σώθηκε στο " + write_path)
 else:
  print("Δεν βρέθηκαν αλλαγές δεν σώζω αρχείο.")
 return(success)

try :
 test_run, attempt, e, retries, wait, headers, cookies, office_run, convert_xl = initialize()
 get_start_time()
 os.system("title " + "Ξεκινήσαμε. Άνοιγμα αρχείων...")
 set_read_files()
 os.system("title " + "Επιλογές...")
 set_sheets()
 set_write_files(sheet_name)
 # sys.exit()
 os.system("title " + "Αρχίζει ο chrome...")
 chrome_options, driver, chrome_drv = chrome_init()
 
 for r in range(row_index + 1, ac_row + 1):
  cell_value = str(sheet.cell(row = r, column = col_index).value)
  comp_list = []
  if cell_value.strip() == "" or cell_value == None or cell_value == "None":
   break
  else:
   current_row = r - 1
   last_row = ac_row - 1
   if current_row < last_row - 1:
    print_text = "Νο.: " + str(current_row) + ". Απομένουν: " + str(ac_row - r) + "/" + str(last_row) + " γραμμές."
   elif current_row == last_row - 1:
    # print_text = "Απομένει: " + str(ac_row - r) + "/" + str(ac_row - 1) + " γραμμή."
    print_text = "Νο.: " + str(current_row) + ". Απομένει 1 γραμμή."
   else:
    print_text = "Τελευταία γραμμή."
   print(print_text)
   comp_name, comp_link, was_price, st_price, st_link, pub_price, pub_link, el_price, el_link, kot_price, kot_link, bio_price, bio_link, si_price, si_link, cus_price, cus_link,comp_wins  = comp_init()
   eshop_code = cell_value.strip()
   os.system("title " + print_text + " - " + eshop_code)
   cy_url = "https://www.e-shop.cy/product?id=" + eshop_code
   gr_url = "https://www.e-shop.gr/s/" + eshop_code
   cy_soup = load_soup(cy_url, wait, retries)
   cy_price, cy_title = get_cy_price(cy_soup)
   
   print("Κωδικός:  " + eshop_code)
   print("Τίτλος:   " + cy_title)
   print("CY Price: " + str(cy_price))
   if cy_price == "ΕΞΑΝΤΛΗΜΕΝΟ":
    print("------------------")
    print("")
    continue
   
   gr_soup = load_soup(gr_url, wait, retries)
   gr_price, gr_title = get_gr_price(gr_soup)
   print("GR Price: " + str(gr_price))
   print("comp_wins: " + str(comp_wins))
   print("type(gr_price)" + str(type(gr_price)))
   print("type(cy_price)" + str(type(cy_price)))
   if type(gr_price) == float and type(cy_price) == float:
    print("gr_price >= cy_price = " + str(gr_price >= cy_price))
   print(">--<")
   
   if type(gr_price) == float and type(cy_price) == float and gr_price >= cy_price:
    comp_wins = True
   else:
    comp_wins = False
   
   print("comp_wins: " + str(comp_wins))

   comp_list.append(eshop_code)
   comp_list.append(cy_title)
   comp_list.append(cy_price)
   comp_list.append(gr_price)

   for c in range(1, ac_col):
    try:
     comp_name = sheet.cell(row = 1, column = c).value
     comp_link = sheet.cell(row = r, column = c).hyperlink.target
     comp_value = sheet.cell(row = r, column = c).value
     # print("comp_name: " + comp_name + ", comp_link: " + comp_link + ", comp_value: " + comp_value)
     if comp_value.find("€") >= 0:
      try:
       # print("Βρήκαμε ευρώ...")
       xl_price = float(comp_value.replace("€", "").replace(",", ".").strip())
       # print("xl_price: " + str(xl_price))
       was_price = str(xl_price)
       # print("was_price: " + was_price)
      except Exception as exc:
       print("Exception: " + str(exc))
       was_price = "-"
     print(comp_name + " link: " + comp_link)
     # print("row: " + str(r))
     # print("column: " + str(c))
     
     if comp_link.find("stephanis") >= 0:
      st_price = st_stuff(comp_link)
      st_link = comp_link
      if st_price < cy_price:
       print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
       comp_wins = True
      was_is(was_price, st_price)
     elif comp_link.find("public") >= 0:
      if chrome_drv == True:
       pub_price = pub_stuff(comp_link)
       pub_link = comp_link
       if pub_price < cy_price:
        print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
        comp_wins = True
       was_is(was_price, pub_price)
      else:
       print("Chrome driver error. Skipping...")
     elif comp_link.find("electroline") >= 0:
      el_price = el_stuff(comp_link)
      el_link = comp_link
      if el_price < cy_price:
       print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
       comp_wins = True
      was_is(was_price, el_price)
     elif comp_link.find("kotsovolos") >= 0:
      kot_price = kot_stuff(comp_link)
      kot_link = comp_link
      if kot_price < cy_price:
       print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
       comp_wins = True
      was_is(was_price, kot_price)
     elif comp_link.find("bionic") >= 0:
      if chrome_drv == True:
       bio_price = bio_stuff(comp_link)
       bio_link = comp_link
       if bio_price < cy_price:
        print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
        comp_wins = True
       was_is(was_price, bio_price)
      else:
       print("Chrome driver error. Skipping...")
     elif comp_link.find("singular") >= 0:
      si_price = sin_stuff(comp_link)
      si_link = comp_link
      if si_price < cy_price:
       print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
       comp_wins = True
      was_is(was_price, si_price)
     elif comp_link.find("custompc") >= 0:
      cus_price = cus_stuff(comp_link)
      cus_link = comp_link
      if cus_price < cy_price:
       print("ΜΙΚΡΟΤΕΡΗ ΤΗΝ ΕΧΕΙΣ")
       comp_wins = True
      was_is(was_price, cus_price)
     # print("")
    except:
     pass
   print("------------------")
   print("")
   # input()
   if comp_wins == True:
    comp_list.append(st_price)
    comp_list.append(st_link)
    comp_list.append(pub_price)
    comp_list.append(pub_link)
    comp_list.append(el_price)
    comp_list.append(el_link)
    comp_list.append(kot_price)
    comp_list.append(kot_link)
    comp_list.append(bio_price)
    comp_list.append(bio_link)
    comp_list.append(si_price)
    comp_list.append(si_link)
    comp_list.append(cus_price)
    comp_list.append(cus_link)
    e = write_results(e, comp_list)
   else:
    continue
except KeyboardInterrupt :
 print("")
 print("OK κατάλαβα. Διαλλειματάκι... ")
 print("")
except Exception as exc :
 print("Άλα της, μόλις πέσαμε πάνω στο exception:")
 print(str(exc))
 print("")
 
# if e > 2 :
#  write_it_down(write_file, alt_write_file, e, sheet_name)
# else :
#  print("Δεν βρέθηκαν αλλαγές δεν σώζω αρχείο.")

try:
 success = write_it_down(e, sheet_index, sheet_name)
 print("")
 print("Bye bye Chrome Driver...")
 driver.quit()
 print("")
 print(sheet_name + " - " + str(sheet_index) + " τέλος.")
 input()
except Exception as exc:
 exception_type, exception_object, exception_traceback = sys.exc_info()
 filename = exception_traceback.tb_frame.f_code.co_filename
 line_number = exception_traceback.tb_lineno
 print("Exception: " + str(exc))
 print("Exception type: ", exception_type)
 print("File name: ", filename)
 print("Line number: ", line_number)
 input()
 sys.exit(0)
